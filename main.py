from fastapi import FastAPI, Depends, HTTPException, UploadFile, File
from fastapi.staticfiles import StaticFiles
from fastapi.templating import Jinja2Templates
from fastapi.requests import Request
from fastapi.responses import HTMLResponse
from sqlalchemy.orm import Session
import uvicorn
import json
from database import engine, get_db, Base
import models, schemas, crud, auth

Base.metadata.create_all(bind=engine)

# Create indexes for fast dashboard queries
from sqlalchemy import text as _text
text = _text
with engine.connect() as _conn:
    for _idx, _col in [
        ("idx_sales_year",     "year"),
        ("idx_sales_month",    "month"),
        ("idx_sales_portal",   "portal"),
        ("idx_sales_branch",   "branch"),
        ("idx_sales_category", "category"),
        ("idx_sales_type",     "type"),
    ]:
        try:
            _conn.execute(_text(f"CREATE INDEX IF NOT EXISTS {_idx} ON sales ({_col})"))
        except: pass
    # Composite index for the most common dashboard query pattern
    try:
        _conn.execute(_text("CREATE INDEX IF NOT EXISTS idx_sales_dash ON sales (year, month, portal, branch, category, type)"))
    except: pass
    # Normalize month to UPPERCASE so UPPER() isn't needed at query time
    try:
        _conn.execute(_text("UPDATE sales SET month = UPPER(TRIM(month)) WHERE month != UPPER(TRIM(month))"))
    except: pass
    # Also normalize portal_sales and portal_files month to UPPERCASE
    try:
        _conn.execute(_text("UPDATE portal_sales SET month = UPPER(TRIM(month)) WHERE month IS NOT NULL AND month != UPPER(TRIM(month))"))
    except: pass
    try:
        _conn.execute(_text("UPDATE portal_files SET month = UPPER(TRIM(month)) WHERE month IS NOT NULL AND month != UPPER(TRIM(month))"))
    except: pass
    _conn.commit()

# ── physical_returns migration: drop old table if columns missing ─────────────
with engine.connect() as _mig:
    try:
        _mig.execute(_text("SELECT date FROM physical_returns LIMIT 1"))
    except Exception:
        try:
            _mig.execute(_text("DROP TABLE IF EXISTS physical_returns"))
            _mig.commit()
        except: pass
        Base.metadata.create_all(bind=engine, tables=[models.PhysicalReturn.__table__])

app = FastAPI(title="Rajnandini ERP")
app.mount("/static", StaticFiles(directory="app/static"), name="static")
templates = Jinja2Templates(directory="app/templates")

@app.post("/api/auth/login")
def login(data: schemas.LoginRequest, db: Session = Depends(get_db)):
    user = crud.authenticate_user(db, data.username, data.password)
    if not user:
        raise HTTPException(status_code=401, detail="Invalid credentials")
    token = auth.create_token({"sub": user.username, "role": user.role})
    return {"token": token, "user": {"username": user.username, "role": user.role, "name": user.name}}

@app.get("/api/company")
def get_company(db: Session = Depends(get_db), current_user=Depends(auth.get_current_user)):
    return crud.get_company(db)

@app.post("/api/company")
def save_company(data: schemas.CompanySchema, db: Session = Depends(get_db), current_user=Depends(auth.get_current_user)):
    if current_user.role != "admin":
        raise HTTPException(403, "Admin only")
    return crud.save_company(db, data)

@app.get("/api/users")
def get_users(db: Session = Depends(get_db), current_user=Depends(auth.get_current_user)):
    if current_user.role != "admin":
        raise HTTPException(403, "Admin only")
    return crud.get_users(db)

@app.post("/api/users")
def create_user(data: schemas.UserCreate, db: Session = Depends(get_db), current_user=Depends(auth.get_current_user)):
    if current_user.role != "admin":
        raise HTTPException(403, "Admin only")
    return crud.create_user(db, data)

@app.put("/api/users/{user_id}")
def update_user(user_id: int, data: schemas.UserUpdate, db: Session = Depends(get_db), current_user=Depends(auth.get_current_user)):
    if current_user.role != "admin":
        raise HTTPException(403, "Admin only")
    return crud.update_user(db, user_id, data)

@app.delete("/api/users/{user_id}")
def delete_user(user_id: int, db: Session = Depends(get_db), current_user=Depends(auth.get_current_user)):
    if current_user.role != "admin":
        raise HTTPException(403, "Admin only")
    return crud.delete_user(db, user_id)


@app.get("/api/sales")
def get_sales(db: Session = Depends(get_db), current_user=Depends(auth.get_current_user)):
    return crud.get_sales(db)

@app.get("/api/sales/types")
def get_sale_types(db: Session = Depends(get_db), current_user=Depends(auth.get_current_user)):
    """Debug: see all distinct type values stored in DB"""
    from sqlalchemy import text
    rows = db.execute(text("SELECT LOWER(TRIM(type)) as t, COUNT(*) as cnt FROM sales GROUP BY t ORDER BY cnt DESC")).fetchall()
    return [{"type": r[0], "count": r[1]} for r in rows]


# ── Static filter cache (portals/branches only — years fetched fresh) ──────────
_filter_cache = {}

def _get_static_filters(db):
    if _filter_cache:
        return _filter_cache
    from sqlalchemy import text
    def distinct(col):
        r = db.execute(text(f"SELECT DISTINCT {col} FROM sales WHERE {col} IS NOT NULL AND {col} != '' ORDER BY {col}")).fetchall()
        return [row[0] for row in r]
    _filter_cache['portals']  = distinct("portal")
    _filter_cache['branches'] = distinct("branch")
    return _filter_cache

@app.post("/api/sales/clear-cache")
def clear_filter_cache(db: Session = Depends(get_db), current_user=Depends(auth.get_current_user)):
    _filter_cache.clear()
    db.expire_all()
    return {"ok": True}

@app.get("/api/sales/dashboard")
def sales_dashboard(
    year: str = None, month: str = None,
    portal: str = None, branch: str = None, category: str = None,
    db: Session = Depends(get_db), current_user=Depends(auth.get_current_user)
):
    from sqlalchemy import text

    # Build WHERE — month stored as-is, compare UPPER only if needed
    conditions = []
    params = {}
    if year     and year     != "ALL": conditions.append("year = :year");         params["year"]     = int(year)
    if month    and month    != "ALL": conditions.append("month = :month");        params["month"]    = month.upper().strip()
    if portal   and portal   != "ALL": conditions.append("portal = :portal");     params["portal"]   = portal
    if branch   and branch   != "ALL": conditions.append("branch = :branch");     params["branch"]   = branch
    if category and category != "ALL": conditions.append("category = :category"); params["category"] = category
    where = ("WHERE " + " AND ".join(conditions)) if conditions else ""

    # Main aggregation query
    sql = text(f"""
        SELECT
            COALESCE(category,'Unknown') as cat,
            LOWER(TRIM(type)) as typ,
            SUM(quantity) as qty,
            SUM(invoice_amount) as amt
        FROM sales
        {where}
        GROUP BY cat, typ
    """)
    rows = db.execute(sql, params).fetchall()

    SALE_TYPES   = {'sale', 'forward'}
    RETURN_TYPES = {'return', 'rto', 'return to origin'}
    CANCEL_TYPES = {'cancel', 'cancelled', 'canceled'}

    sale_qty=sale_amt=ret_qty=ret_amt=can_qty=can_amt=0.0
    cat_map = {}
    for cat, typ, qty, amt in rows:
        qty = qty or 0; amt = amt or 0
        is_s = typ in SALE_TYPES
        is_r = typ in RETURN_TYPES
        is_c = typ in CANCEL_TYPES
        if is_s: sale_qty+=qty; sale_amt+=amt
        if is_r: ret_qty+=qty;  ret_amt+=amt
        if is_c: can_qty+=qty;  can_amt+=amt
        if cat not in cat_map:
            cat_map[cat] = [0,0.0,0,0.0]
        if is_s: cat_map[cat][0]+=qty; cat_map[cat][1]+=amt
        if is_r: cat_map[cat][2]+=qty; cat_map[cat][3]+=amt

    net_qty = sale_qty - ret_qty
    net_amt = sale_amt - ret_amt

    cat_rows = []
    for cat, (sq,sa,rq,ra) in sorted(cat_map.items(), key=lambda x: -x[1][0]):
        nq=sq-rq; na=sa-ra
        cat_rows.append({
            "category": cat,
            "sale_qty": int(sq), "sale_amt": round(sa,2),
            "ret_qty":  int(rq), "ret_amt":  round(ra,2),
            "net_qty":  int(nq), "net_amt":  round(na,2),
            "avg_sp":   round(na/nq,0) if nq else 0,
            "pct":      round(sq/sale_qty*100,2) if sale_qty else 0,
        })

    # Months — single fast query scoped to current year/portal/branch (no month filter)
    MONTH_ORDER = ["JANUARY","FEBRUARY","MARCH","APRIL","MAY","JUNE",
                   "JULY","AUGUST","SEPTEMBER","OCTOBER","NOVEMBER","DECEMBER"]
    cond_nm = [c for c in conditions if "month" not in c]
    where_nm = ("WHERE " + " AND ".join(cond_nm) + " AND month IS NOT NULL AND month != ''") if cond_nm else "WHERE month IS NOT NULL AND month != ''"
    params_nm = {k:v for k,v in params.items() if k != "month"}
    raw_months = db.execute(text(f"SELECT DISTINCT month FROM sales {where_nm}"), params_nm).fetchall()
    db_months = sorted({r[0] for r in raw_months if r[0]}, key=lambda m: MONTH_ORDER.index(m) if m in MONTH_ORDER else 99)

    # Static filters from cache (portals/branches only)
    sf = _get_static_filters(db)

    # Years — always fresh from DB (reflects actual data)
    raw_years = db.execute(text("SELECT DISTINCT year FROM sales WHERE year IS NOT NULL ORDER BY year DESC")).fetchall()
    db_years = [r[0] for r in raw_years if r[0]]

    return {
        "summary": {
            "sale_qty": int(sale_qty), "sale_amt": round(sale_amt,2),
            "ret_qty":  int(ret_qty),  "ret_amt":  round(ret_amt,2),
            "net_qty":  int(net_qty),  "net_amt":  round(net_amt,2),
            "avg_sp":   round(net_amt/net_qty,0) if net_qty else 0,
            "can_qty":  int(can_qty),  "can_amt":  round(can_amt,2),
        },
        "by_category": cat_rows,
        "filters": {
            "years":    db_years,
            "months":   db_months,
            "portals":  sf['portals'],
            "branches": sf['branches'],
        }
    }


@app.get("/api/sales/imported-files")
def get_imported_files(db: Session = Depends(get_db), current_user=Depends(auth.get_current_user)):
    import os, glob
    from sqlalchemy import text
    # Get all files currently in the folder
    folder = r"C:\Users\LENOVO\OneDrive\Rajnandini\01 IMP_Do Not Edit  Delete\04 Sales Data"
    folder_files = set()
    if os.path.exists(folder):
        for ext in ["*.xlsx","*.xls","*.csv"]:
            for f in glob.glob(os.path.join(folder,"**",ext), recursive=True):
                folder_files.add(os.path.basename(f))

    # Get all imported file names
    rows = db.execute(text("""
        SELECT file_name, COUNT(*) as rows, MAX(uploaded_at) as last_import
        FROM sales WHERE file_name IS NOT NULL AND file_name != ''
        GROUP BY file_name ORDER BY last_import DESC
    """)).fetchall()

    deleted_any = False
    result = []
    for r in rows:
        name = r[0]
        if name not in folder_files:
            # File deleted from folder — purge its data from DB
            db.execute(text("DELETE FROM sales WHERE file_name = :name"), {"name": name})
            deleted_any = True
        else:
            # Get file modified time
            for ext in ["*.xlsx","*.xls","*.csv"]:
                for f in glob.glob(os.path.join(folder,"**",ext), recursive=True):
                    if os.path.basename(f) == name:
                        result.append({
                            "name": name,
                            "rows": r[1],
                            "imported_at": r[2],
                            "file_modified": round(os.stat(f).st_mtime),
                        })
                        break

    if deleted_any:
        db.commit()
        _filter_cache.clear()  # reset cache since data changed

    return result

@app.get("/api/sales/folder")
def scan_sales_folder(current_user=Depends(auth.get_current_user)):
    import os, glob
    folder = r"C:\Users\LENOVO\OneDrive\Rajnandini\01 IMP_Do Not Edit  Delete\04 Sales Data"
    if not os.path.exists(folder):
        raise HTTPException(404, detail=f"Folder not found: {folder}")
    files = []
    for ext in ["*.xlsx", "*.xls", "*.csv"]:
        for f in glob.glob(os.path.join(folder, "**", ext), recursive=True):
            stat = os.stat(f)
            files.append({
                "name": os.path.basename(f),
                "path": f,
                "size_kb": round(stat.st_size / 1024, 1),
                "modified": stat.st_mtime,
            })
    files.sort(key=lambda x: x["modified"], reverse=True)
    return {"folder": folder, "files": files}

@app.post("/api/sales/process-file")
async def process_sale_file(request: Request, db: Session = Depends(get_db), current_user=Depends(auth.get_current_user)):
    import os
    data = await request.json()
    file_path = data.get("path", "")
    if not os.path.exists(file_path):
        raise HTTPException(404, detail="File not found")
    try:
        import openpyxl
        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        ws = wb.active
        rows_data = list(ws.iter_rows(values_only=True))
        wb.close()
        if not rows_data:
            return {"inserted": 0, "skipped": 0}
        headers = [str(h).strip().lower().replace(" ", "_") if h else "" for h in rows_data[0]]
        result = crud.bulk_insert_sales(db, [
            dict(zip(headers, row)) for row in rows_data[1:] if any(row)
        ], current_user.username, os.path.basename(file_path))
        return result
    except Exception as e:
        raise HTTPException(500, detail=str(e))

@app.get("/api/sales/debug")
def debug_sales(db: Session = Depends(get_db), current_user=Depends(auth.get_current_user)):
    from sqlalchemy import text
    years = db.execute(text("SELECT DISTINCT year, typeof(year) FROM sales LIMIT 5")).fetchall()
    months = db.execute(text("SELECT DISTINCT month, year FROM sales ORDER BY year DESC LIMIT 20")).fetchall()
    return {
        "years_sample": [{"year": r[0], "type": r[1]} for r in years],
        "months_sample": [{"month": r[0], "year": r[1]} for r in months],
    }

@app.get("/api/sales/process-file-stream")
async def process_sale_file_stream(path: str, token: str, db: Session = Depends(get_db)):
    import os, json
    from fastapi.responses import StreamingResponse
    from jose import jwt, JWTError
    SECRET_KEY = "rajnandini-fashion-secret-key-2024"
    try:
        payload = jwt.decode(token, SECRET_KEY, algorithms=["HS256"])
        username = payload.get("sub")
        if not username:
            raise HTTPException(401, "Invalid token")
    except JWTError:
        raise HTTPException(401, "Invalid token")

    if not os.path.exists(path):
        raise HTTPException(404, detail="File not found")

    async def event_stream():
        try:
            import openpyxl, asyncio
            # Yield immediately so browser sees the connection and shows bar
            _d = {'stage':'reading','pct':2}
            yield f"data: {json.dumps(_d)}\n\n"
            await asyncio.sleep(0)  # flush to client

            # Run heavy file read in thread so event loop stays responsive
            import concurrent.futures
            loop = asyncio.get_event_loop()
            def read_file():
                wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
                ws = wb.active
                data = list(ws.iter_rows(values_only=True))
                wb.close()
                return data

            _d = {'stage':'reading','pct':5}
            yield f"data: {json.dumps(_d)}\n\n"
            await asyncio.sleep(0)

            with concurrent.futures.ThreadPoolExecutor() as pool:
                rows_data = await loop.run_in_executor(pool, read_file)

            total = max(len(rows_data) - 1, 1)
            _d = {'stage':'parsing','pct':15,'total':total}
            yield f"data: {json.dumps(_d)}\n\n"
            await asyncio.sleep(0)

            if not rows_data:
                _d = {'stage':'done','pct':100,'inserted':0,'skipped':0}
                yield f"data: {json.dumps(_d)}\n\n"
                return

            headers = [str(h).strip().lower().replace(" ", "_") if h else "" for h in rows_data[0]]
            data_rows = [dict(zip(headers, row)) for row in rows_data[1:] if any(row)]
            _d = {'stage':'parsing','pct':20,'total':len(data_rows)}
            yield f"data: {json.dumps(_d)}\n\n"
            await asyncio.sleep(0)

            CHUNK = 500
            inserted = 0; skipped = 0
            fname = os.path.basename(path)
            # Delete existing rows ONCE before inserting chunks
            from sqlalchemy import text as _t
            db.execute(_t("DELETE FROM sales WHERE file_name = :fn"), {"fn": fname})
            db.commit()
            for i in range(0, len(data_rows), CHUNK):
                chunk = data_rows[i:i+CHUNK]
                res = crud.bulk_insert_sales(db, chunk, username, fname, delete_first=False)
                inserted += res.get("inserted", 0)
                skipped  += res.get("skipped", 0)
                pct = 20 + int(((i + len(chunk)) / len(data_rows)) * 78)
                _d = {'stage':'inserting','pct':pct,'inserted':inserted,'skipped':skipped}
                yield f"data: {json.dumps(_d)}\n\n"
                await asyncio.sleep(0)  # flush chunk progress to browser

            _d = {'stage':'done','pct':100,'inserted':inserted,'skipped':skipped}
            yield f"data: {json.dumps(_d)}\n\n"
        except Exception as e:
            _d = {'stage':'error','msg':str(e)}
            yield f"data: {json.dumps(_d)}\n\n"

    return StreamingResponse(event_stream(), media_type="text/event-stream",
        headers={"Cache-Control":"no-cache","X-Accel-Buffering":"no"})

@app.post("/api/sales/upload")
async def upload_sales(request: Request, db: Session = Depends(get_db), current_user=Depends(auth.get_current_user)):
    data = await request.json()
    return crud.bulk_insert_sales(db, data["rows"], current_user.username, data.get("file_name", ""))

@app.get("/api/sales/summary")
def sales_summary(db: Session = Depends(get_db), current_user=Depends(auth.get_current_user)):
    return crud.get_sales_summary(db)


# ── PORTAL SALE ───────────────────────────────────────────────────────────────

import os as _os, getpass as _getpass

def get_portal_sale_folder():
    user = _getpass.getuser()
    return f"C:\\Users\\{user}\\OneDrive\\Rajnandini\\01 IMP_Do Not Edit  Delete\\06 Portal Sales Files\\Amazon"

def ensure_portal_sale_tables(db):
    from sqlalchemy import text
    db.execute(text("""
        CREATE TABLE IF NOT EXISTS portal_sale_uploads (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            portal TEXT NOT NULL,
            report_type TEXT,
            file_name TEXT NOT NULL,
            month TEXT,
            year INTEGER,
            account TEXT,
            rows_imported INTEGER DEFAULT 0,
            uploaded_by TEXT,
            uploaded_at DATETIME DEFAULT CURRENT_TIMESTAMP
        )
    """))
    db.execute(text("""
        CREATE TABLE IF NOT EXISTS portal_sales (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            upload_id INTEGER,
            portal TEXT,
            report_type TEXT,
            month TEXT,
            year INTEGER,
            seller_gstin TEXT,
            invoice_number TEXT,
            invoice_date TEXT,
            transaction_type TEXT,
            mapped_type TEXT,
            order_id TEXT,
            shipment_id TEXT,
            shipment_date TEXT,
            order_date TEXT,
            sku TEXT,
            mapped_sku TEXT,
            asin TEXT,
            item_description TEXT,
            quantity INTEGER DEFAULT 0,
            invoice_amount REAL DEFAULT 0,
            tax_exclusive_gross REAL DEFAULT 0,
            ship_from_city TEXT,
            ship_from_state TEXT,
            ship_to_city TEXT,
            ship_to_state TEXT,
            fulfillment_channel TEXT,
            warehouse_id TEXT,
            payment_method TEXT,
            buyer_name TEXT,
            buyer_gstin TEXT
        )
    """))
    db.execute(text("""
        CREATE TABLE IF NOT EXISTS portal_txn_type_map (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            portal TEXT NOT NULL,
            transaction_type TEXT NOT NULL,
            mapped_type TEXT NOT NULL,
            UNIQUE(portal, transaction_type)
        )
    """))
    db.commit()

def _detect_month_year(filename):
    import re
    MONTHS = ["JANUARY","FEBRUARY","MARCH","APRIL","MAY","JUNE",
              "JULY","AUGUST","SEPTEMBER","OCTOBER","NOVEMBER","DECEMBER"]
    fn = filename.upper()
    month = next((m for m in MONTHS if m in fn), None)
    ym = re.search(r'20\d{2}', fn)
    year = int(ym.group()) if ym else None
    return month, year

def _gcol(row, *names):
    for n in names:
        nl = n.lower().strip()
        for k, v in row.items():
            if k.lower().strip() == nl:
                return str(v).strip() if v is not None else ""
    return ""

@app.get("/api/portal-sales/folder")
def ps_folder(current_user=Depends(auth.get_current_user)):
    folder = get_portal_sale_folder()
    result = {}
    if not _os.path.exists(folder):
        return {"folder": folder, "exists": False, "portals": {}}
    # Amazon folder — scan subfolders (month/year names) recursively for CSV files
    files_all = []
    for subfolder in _os.listdir(folder):
        sp = _os.path.join(folder, subfolder)
        if _os.path.isdir(sp):
            for f in _os.listdir(sp):
                if f.lower().endswith('.csv'):
                    fp = _os.path.join(sp, f)
                    stat = _os.stat(fp)
                    month, year = _detect_month_year(f)
                    files_all.append({"name": f, "path": fp,
                        "size_kb": round(stat.st_size/1024, 1),
                        "modified": stat.st_mtime, "month": month, "year": year,
                        "subfolder": subfolder})
        elif subfolder.lower().endswith('.csv'):
            fp = _os.path.join(folder, subfolder)
            stat = _os.stat(fp)
            month, year = _detect_month_year(subfolder)
            files_all.append({"name": subfolder, "path": fp,
                "size_kb": round(stat.st_size/1024, 1),
                "modified": stat.st_mtime, "month": month, "year": year,
                "subfolder": ""})
    files_all.sort(key=lambda x: x["modified"], reverse=True)
    result["Amazon"] = files_all
    return {"folder": folder, "exists": True, "portals": result}

@app.get("/api/portal-sales/uploads")
def ps_uploads(portal: str = None, db: Session = Depends(get_db),
               current_user=Depends(auth.get_current_user)):
    from sqlalchemy import text
    ensure_portal_sale_tables(db)
    q = "SELECT * FROM portal_sale_uploads"
    params = {}
    if portal and portal != "ALL":
        q += " WHERE portal = :portal"; params["portal"] = portal
    rows = db.execute(text(q + " ORDER BY uploaded_at DESC"), params).fetchall()
    return [dict(r._mapping) for r in rows]

@app.get("/api/portal-sales/txn-map")
def ps_txn_map(portal: str = None, db: Session = Depends(get_db),
               current_user=Depends(auth.get_current_user)):
    from sqlalchemy import text
    ensure_portal_sale_tables(db)
    q = "SELECT * FROM portal_txn_type_map"
    params = {}
    if portal:
        q += " WHERE portal=:portal"; params["portal"] = portal
    rows = db.execute(text(q), params).fetchall()
    return [dict(r._mapping) for r in rows]

@app.post("/api/portal-sales/txn-map")
async def ps_save_txn_map(request: Request, db: Session = Depends(get_db),
                           current_user=Depends(auth.get_current_user)):
    from sqlalchemy import text
    ensure_portal_sale_tables(db)
    d = await request.json()
    p, t, m = d["portal"], d["transaction_type"], d["mapped_type"]
    db.execute(text("""
        INSERT INTO portal_txn_type_map (portal,transaction_type,mapped_type)
        VALUES (:p,:t,:m)
        ON CONFLICT(portal,transaction_type) DO UPDATE SET mapped_type=:m
    """), {"p":p,"t":t,"m":m})
    db.execute(text("""
        UPDATE portal_sales SET mapped_type=:m
        WHERE portal=:p AND LOWER(TRIM(transaction_type))=LOWER(TRIM(:t))
    """), {"p":p,"t":t,"m":m})
    db.commit()
    return {"ok": True}

@app.get("/api/portal-sales/import-stream")
async def ps_import_stream(path: str, portal: str, token: str,
                            db: Session = Depends(get_db)):
    import csv, json, asyncio, re
    from fastapi.responses import StreamingResponse
    from jose import jwt, JWTError
    from sqlalchemy import text
    SECRET_KEY = "rajnandini-fashion-secret-key-2024"
    try:
        payload = jwt.decode(token, SECRET_KEY, algorithms=["HS256"])
        username = payload.get("sub")
        if not username: raise HTTPException(401)
    except: raise HTTPException(401)
    ensure_portal_sale_tables(db)

    async def stream():
        try:
            if not _os.path.exists(path):
                yield f"data: {json.dumps({'stage':'error','msg':'File not found'})}\n\n"; return
            yield f"data: {json.dumps({'stage':'reading','pct':5})}\n\n"
            await asyncio.sleep(0)
            fname = _os.path.basename(path)
            month, year = _detect_month_year(fname)
            fn_upper = fname.upper()
            report_type = "B2B" if "B2B" in fn_upper else ("B2C" if "B2C" in fn_upper else "MTR")
            acc_m = re.search(r'[A-Z0-9]{12,}', fname.upper())
            account = acc_m.group() if acc_m else fname

            with open(path, encoding='utf-8-sig') as f:
                reader = csv.DictReader(f)
                raw_rows = list(reader)
            total = len(raw_rows)
            yield f"data: {json.dumps({'stage':'parsing','pct':15,'total':total})}\n\n"
            await asyncio.sleep(0)

            txn_map = {r[0].lower().strip(): r[1] for r in
                db.execute(text("SELECT transaction_type,mapped_type FROM portal_txn_type_map WHERE portal=:p"), {"p":portal}).fetchall()}
            sku_map = {r[0].strip(): r[1].strip() for r in
                db.execute(text("SELECT portal_sku,item_master_sku FROM sku_mappings WHERE portal=:p AND item_master_sku IS NOT NULL"), {"p":portal}).fetchall()}

            existing = db.execute(text("SELECT id FROM portal_sale_uploads WHERE file_name=:fn AND portal=:p"), {"fn":fname,"p":portal}).fetchone()
            if existing:
                db.execute(text("DELETE FROM portal_sales WHERE upload_id=:uid"), {"uid":existing[0]})
                db.execute(text("DELETE FROM portal_sale_uploads WHERE id=:uid"), {"uid":existing[0]})
                db.commit()

            db.execute(text("""INSERT INTO portal_sale_uploads (portal,report_type,file_name,month,year,account,uploaded_by)
                VALUES (:po,:rt,:fn,:mo,:yr,:ac,:ub)"""),
                {"po":portal,"rt":report_type,"fn":fname,"mo":month,"yr":year,"ac":account,"ub":username})
            db.commit()
            upload_id = db.execute(text("SELECT last_insert_rowid()")).scalar()

            unique_txn = set()
            inserted = 0
            CHUNK = 300
            batch = []

            def make_row(row):
                txn = _gcol(row, "Transaction Type")
                sku = _gcol(row, "Sku", "SKU")
                unique_txn.add(txn)
                try: qty = int(float(_gcol(row,"Quantity") or 0))
                except: qty = 0
                try: inv = float(_gcol(row,"Invoice Amount") or 0)
                except: inv = 0
                try: teg = float(_gcol(row,"Tax Exclusive Gross") or 0)
                except: teg = 0
                desc = _gcol(row,"Item Description")
                return {"upload_id":upload_id,"portal":portal,"report_type":report_type,
                    "month":(month or "").upper().strip() or None,"year":year,"seller_gstin":_gcol(row,"Seller Gstin"),
                    "invoice_number":_gcol(row,"Invoice Number"),"invoice_date":_gcol(row,"Invoice Date"),
                    "transaction_type":txn,"mapped_type":txn_map.get(txn.lower().strip(),""),
                    "order_id":_gcol(row,"Order Id"),"shipment_id":_gcol(row,"Shipment Id"),
                    "shipment_date":_gcol(row,"Shipment Date"),"order_date":_gcol(row,"Order Date"),
                    "sku":sku,"mapped_sku":sku_map.get(sku,""),"asin":_gcol(row,"Asin"),
                    "item_description":desc[:200] if desc else "",
                    "quantity":qty,"invoice_amount":inv,"tax_exclusive_gross":teg,
                    "ship_from_city":_gcol(row,"Ship From City"),"ship_from_state":_gcol(row,"Ship From State"),
                    "ship_to_city":_gcol(row,"Ship To City"),"ship_to_state":_gcol(row,"Ship To State"),
                    "fulfillment_channel":_gcol(row,"Fulfillment Channel"),"warehouse_id":_gcol(row,"Warehouse Id"),
                    "payment_method":_gcol(row,"Payment Method Code"),
                    "buyer_name":_gcol(row,"Buyer Name"),"buyer_gstin":_gcol(row,"Customer Bill To Gstid")}

            INSERT_SQL = text("""INSERT INTO portal_sales
                (upload_id,portal,report_type,month,year,seller_gstin,invoice_number,invoice_date,
                transaction_type,mapped_type,order_id,shipment_id,shipment_date,order_date,
                sku,mapped_sku,asin,item_description,quantity,invoice_amount,tax_exclusive_gross,
                ship_from_city,ship_from_state,ship_to_city,ship_to_state,
                fulfillment_channel,warehouse_id,payment_method,buyer_name,buyer_gstin)
                VALUES
                (:upload_id,:portal,:report_type,:month,:year,:seller_gstin,:invoice_number,:invoice_date,
                :transaction_type,:mapped_type,:order_id,:shipment_id,:shipment_date,:order_date,
                :sku,:mapped_sku,:asin,:item_description,:quantity,:invoice_amount,:tax_exclusive_gross,
                :ship_from_city,:ship_from_state,:ship_to_city,:ship_to_state,
                :fulfillment_channel,:warehouse_id,:payment_method,:buyer_name,:buyer_gstin)""")

            for i, row in enumerate(raw_rows):
                batch.append(make_row(row))
                if len(batch) >= CHUNK:
                    db.execute(INSERT_SQL, batch); db.commit()
                    inserted += len(batch); batch = []
                    pct = 15 + int((i/total)*80)
                    yield f"data: {json.dumps({'stage':'inserting','pct':pct,'inserted':inserted})}\n\n"
                    await asyncio.sleep(0)
            if batch:
                db.execute(INSERT_SQL, batch); db.commit(); inserted += len(batch)

            db.execute(text("UPDATE portal_sale_uploads SET rows_imported=:r WHERE id=:uid"), {"r":inserted,"uid":upload_id})
            db.commit()
            yield f"data: {json.dumps({'stage':'done','pct':100,'inserted':inserted,'unique_txn':list(unique_txn),'upload_id':upload_id})}\n\n"
        except Exception as e:
            import traceback
            yield f"data: {json.dumps({'stage':'error','msg':str(e)})}\n\n"

    return StreamingResponse(stream(), media_type="text/event-stream",
        headers={"Cache-Control":"no-cache","X-Accel-Buffering":"no"})

@app.get("/api/portal-sales/dashboard")
def ps_dashboard(portal: str = None, year: str = None, month: str = None,
                 db: Session = Depends(get_db), current_user=Depends(auth.get_current_user)):
    from sqlalchemy import text
    ensure_portal_sale_tables(db)
    conds = []; params = {}
    if portal and portal != "ALL": conds.append("portal=:portal"); params["portal"] = portal
    if year and year != "ALL": conds.append("year=:year"); params["year"] = int(year)
    if month and month != "ALL": conds.append("month=:month"); params["month"] = month
    where = ("WHERE " + " AND ".join(conds)) if conds else ""

    rows = db.execute(text(f"SELECT transaction_type, mapped_type, SUM(quantity), SUM(invoice_amount) FROM portal_sales {where} GROUP BY transaction_type, mapped_type"), params).fetchall()
    txn_summary = {}
    for r in rows:
        txn_summary[r[0]] = {"mapped_type": r[1] or "", "qty": int(r[2] or 0), "amt": round(r[3] or 0, 2)}

    portals = [r[0] for r in db.execute(text("SELECT DISTINCT portal FROM portal_sale_uploads ORDER BY portal")).fetchall()]
    years = [r[0] for r in db.execute(text("SELECT DISTINCT year FROM portal_sales WHERE year IS NOT NULL ORDER BY year DESC")).fetchall()]
    MONTH_ORDER = ["JANUARY","FEBRUARY","MARCH","APRIL","MAY","JUNE","JULY","AUGUST","SEPTEMBER","OCTOBER","NOVEMBER","DECEMBER"]
    months_raw = [r[0] for r in db.execute(text("SELECT DISTINCT month FROM portal_sales WHERE month IS NOT NULL")).fetchall() if r[0]]
    months = sorted(months_raw, key=lambda m: MONTH_ORDER.index(m) if m in MONTH_ORDER else 99)

    txn_map = {r[0]: r[1] for r in db.execute(text("SELECT transaction_type,mapped_type FROM portal_txn_type_map WHERE portal=:p"), {"p": portal or "Amazon"}).fetchall()}
    all_txn = [r[0] for r in db.execute(text("SELECT DISTINCT transaction_type FROM portal_sales WHERE portal=:p"), {"p": portal or "Amazon"}).fetchall()]

    return {"txn_summary": txn_summary, "txn_map": txn_map, "all_txn": all_txn,
            "filters": {"portals": portals, "years": years, "months": months}}

@app.get("/api/portal-sales/unmapped-skus")
def ps_unmapped_skus(portal: str = "Amazon", db: Session = Depends(get_db),
                     current_user=Depends(auth.get_current_user)):
    from sqlalchemy import text
    ensure_portal_sale_tables(db)
    rows = db.execute(text("""
        SELECT DISTINCT ps.sku, COUNT(*) as cnt FROM portal_sales ps
        LEFT JOIN sku_mappings sm ON ps.sku=sm.portal_sku AND sm.portal=ps.portal
        WHERE ps.portal=:p AND (ps.mapped_sku IS NULL OR ps.mapped_sku='') AND sm.id IS NULL
        GROUP BY ps.sku ORDER BY cnt DESC LIMIT 200
    """), {"p": portal}).fetchall()
    return [{"sku": r[0], "count": r[1]} for r in rows if r[0]]

@app.post("/api/portal-sales/map-sku")
async def ps_map_sku(request: Request, db: Session = Depends(get_db),
                     current_user=Depends(auth.get_current_user)):
    from sqlalchemy import text
    d = await request.json()
    ps, po, im = d["portal_sku"], d["portal"], d["item_master_sku"]
    try:
        db.execute(text("INSERT INTO sku_mappings (portal_sku,portal,item_master_sku) VALUES (:ps,:po,:im) ON CONFLICT(portal_sku,portal) DO UPDATE SET item_master_sku=:im"), {"ps":ps,"po":po,"im":im})
    except:
        db.execute(text("UPDATE sku_mappings SET item_master_sku=:im WHERE portal_sku=:ps AND portal=:po"), {"im":im,"ps":ps,"po":po})
    db.execute(text("UPDATE portal_sales SET mapped_sku=:im WHERE sku=:ps AND portal=:po"), {"im":im,"ps":ps,"po":po})
    db.commit()
    return {"ok": True}


# ── PORTAL SALE NEW API ───────────────────────────────────────────────────────

@app.get("/api/portal-sale/folder-path")
def ps_new_folder_path(current_user=Depends(auth.get_current_user)):
    return {"path": get_portal_sale_folder()}

@app.post("/api/portal-sale/scan")
def ps_new_scan(db: Session = Depends(get_db), current_user=Depends(auth.get_current_user)):
    from sqlalchemy import text
    import csv
    ensure_portal_sale_tables(db)
    # Amazon folder — files live inside month/year subfolders
    amazon_folder = get_portal_sale_folder()
    portal = "Amazon"
    scanned = []
    if not _os.path.exists(amazon_folder):
        return {"scanned": 0, "error": f"Folder not found: {amazon_folder}"}
    # Collect all CSV/XLSX files from subfolders (month/year names) and root level
    all_files = []
    for entry in _os.listdir(amazon_folder):
        ep = _os.path.join(amazon_folder, entry)
        if _os.path.isdir(ep):
            for fname in _os.listdir(ep):
                if fname.lower().endswith(('.csv', '.xlsx', '.xls')):
                    all_files.append((fname, _os.path.join(ep, fname)))
        elif entry.lower().endswith(('.csv', '.xlsx', '.xls')):
            all_files.append((entry, ep))
    for fname, fpath in all_files:
        existing = db.execute(text("SELECT id,status FROM ps_files WHERE filename=:fn AND portal=:p"), {"fn":fname,"p":portal}).fetchone()
        if existing: continue
        try:
            with open(fpath, encoding='utf-8-sig') as f:
                rows = list(csv.reader(f))
            row_count = len(rows) - 1
        except: row_count = 0
        db.execute(text("""INSERT INTO ps_files (portal,filename,filepath,row_count,status,scanned_at)
            VALUES (:po,:fn,:fp,:rc,'pending',datetime('now'))"""),
            {"po":portal,"fn":fname,"fp":fpath,"rc":row_count})
        scanned.append(fname)
    db.commit()
    return {"scanned": len(scanned), "files": scanned}

@app.get("/api/portal-sale/files")
def ps_new_files(db: Session = Depends(get_db), current_user=Depends(auth.get_current_user)):
    from sqlalchemy import text
    ensure_portal_sale_tables(db)
    rows = db.execute(text("SELECT * FROM ps_files ORDER BY portal,filename")).fetchall()
    return [dict(r._mapping) for r in rows]

@app.post("/api/portal-sale/load/{file_id}")
def ps_new_load(file_id: int, db: Session = Depends(get_db), current_user=Depends(auth.get_current_user)):
    import csv
    from sqlalchemy import text
    ensure_portal_sale_tables(db)
    f = db.execute(text("SELECT * FROM ps_files WHERE id=:id"), {"id":file_id}).fetchone()
    if not f: raise HTTPException(404, "File not found")
    fdata = dict(f._mapping)
    fpath = fdata["filepath"]
    portal = fdata["portal"]
    fname = fdata["filename"]
    if not _os.path.exists(fpath): raise HTTPException(404, "CSV file not found on disk")

    # Load txn map and sku map
    txn_map = {r[0].lower().strip(): r[1] for r in
        db.execute(text("SELECT transaction_type,mapped_type FROM portal_txn_type_map WHERE portal=:p"), {"p":portal}).fetchall()}
    sku_map = {r[0].strip(): r[1].strip() for r in
        db.execute(text("SELECT portal_sku,item_master_sku FROM sku_mappings WHERE portal=:p AND item_master_sku IS NOT NULL"), {"p":portal}).fetchall()}

    # Delete existing rows for this file
    db.execute(text("DELETE FROM ps_rows WHERE file_id=:fid"), {"fid":file_id})
    db.commit()

    with open(fpath, encoding='utf-8-sig') as csvf:
        reader = csv.DictReader(csvf)
        raw = list(reader)

    inserted = 0
    batch = []
    INSERT_SQL = text("""INSERT INTO ps_rows
        (file_id,portal,filename,transaction_type,mapped_as,portal_sku,item_master_sku,
        order_id,invoice_number,invoice_date,quantity,invoice_amount,tax_exclusive_gross,
        ship_from_city,ship_from_state,ship_to_city,ship_to_state,fulfillment_channel,
        warehouse_id,seller_gstin,asin,buyer_name,buyer_gstin)
        VALUES
        (:file_id,:portal,:filename,:transaction_type,:mapped_as,:portal_sku,:item_master_sku,
        :order_id,:invoice_number,:invoice_date,:quantity,:invoice_amount,:tax_exclusive_gross,
        :ship_from_city,:ship_from_state,:ship_to_city,:ship_to_state,:fulfillment_channel,
        :warehouse_id,:seller_gstin,:asin,:buyer_name,:buyer_gstin)""")

    for row in raw:
        txn = _gcol(row,"Transaction Type")
        sku = _gcol(row,"Sku","SKU")
        try: qty = int(float(_gcol(row,"Quantity") or 0))
        except: qty = 0
        try: inv = float(_gcol(row,"Invoice Amount") or 0)
        except: inv = 0
        try: teg = float(_gcol(row,"Tax Exclusive Gross") or 0)
        except: teg = 0
        mapped_as = txn_map.get(txn.lower().strip(), "")
        im_sku = sku_map.get(sku, "")
        batch.append({"file_id":file_id,"portal":portal,"filename":fname,
            "transaction_type":txn,"mapped_as":mapped_as,"portal_sku":sku,"item_master_sku":im_sku,
            "order_id":_gcol(row,"Order Id"),"invoice_number":_gcol(row,"Invoice Number"),
            "invoice_date":_gcol(row,"Invoice Date"),"quantity":qty,
            "invoice_amount":inv,"tax_exclusive_gross":teg,
            "ship_from_city":_gcol(row,"Ship From City"),"ship_from_state":_gcol(row,"Ship From State"),
            "ship_to_city":_gcol(row,"Ship To City"),"ship_to_state":_gcol(row,"Ship To State"),
            "fulfillment_channel":_gcol(row,"Fulfillment Channel"),
            "warehouse_id":_gcol(row,"Warehouse Id"),
            "seller_gstin":_gcol(row,"Seller Gstin"),"asin":_gcol(row,"Asin"),
            "buyer_name":_gcol(row,"Buyer Name"),"buyer_gstin":_gcol(row,"Customer Bill To Gstid")})
        if len(batch) >= 500:
            db.execute(INSERT_SQL, batch); db.commit(); inserted += len(batch); batch = []
    if batch:
        db.execute(INSERT_SQL, batch); db.commit(); inserted += len(batch)

    # Register new txn types in mapping table
    txn_types = list({r["transaction_type"] for r in batch if r["transaction_type"]})
    for t in set(_gcol(row,"Transaction Type") for row in raw if _gcol(row,"Transaction Type")):
        db.execute(text("""INSERT OR IGNORE INTO portal_txn_type_map (portal,transaction_type,mapped_type)
            VALUES (:p,:t,'')"""), {"p":portal,"t":t})
    db.commit()

    db.execute(text("UPDATE ps_files SET status='loaded',row_count=:rc WHERE id=:id"), {"rc":inserted,"id":file_id})
    db.commit()
    return {"ok":True,"rows":inserted}

@app.get("/api/portal-sale/mapping")
def ps_new_get_mapping(db: Session = Depends(get_db), current_user=Depends(auth.get_current_user)):
    from sqlalchemy import text
    ensure_portal_sale_tables(db)
    rows = db.execute(text("SELECT portal,transaction_type,mapped_type FROM portal_txn_type_map ORDER BY portal,transaction_type")).fetchall()
    result = {}
    for r in rows:
        if r[0] not in result: result[r[0]] = {}
        result[r[0]][r[1]] = r[2] or ""
    return result

@app.post("/api/portal-sale/mapping")
async def ps_new_save_mapping(request: Request, db: Session = Depends(get_db),
                               current_user=Depends(auth.get_current_user)):
    from sqlalchemy import text
    data = await request.json()  # {portal: {txn: mapped}}
    for portal, types in data.items():
        for txn, mapped in types.items():
            db.execute(text("""INSERT INTO portal_txn_type_map (portal,transaction_type,mapped_type)
                VALUES (:p,:t,:m) ON CONFLICT(portal,transaction_type) DO UPDATE SET mapped_type=:m"""),
                {"p":portal,"t":txn,"m":mapped})
            db.execute(text("UPDATE ps_rows SET mapped_as=:m WHERE portal=:p AND LOWER(TRIM(transaction_type))=LOWER(TRIM(:t))"),
                {"m":mapped,"p":portal,"t":txn})
    db.commit()
    return {"ok":True}

@app.get("/api/portal-sale/data")
def ps_new_data(portal: str = None, mapped_as: str = None, unmapped: int = 0,
                page: int = 1, per_page: int = 50,
                db: Session = Depends(get_db), current_user=Depends(auth.get_current_user)):
    from sqlalchemy import text
    ensure_portal_sale_tables(db)
    conds = []; params = {}
    if portal: conds.append("r.portal=:portal"); params["portal"] = portal
    if mapped_as: conds.append("r.mapped_as=:mapped_as"); params["mapped_as"] = mapped_as
    if unmapped: conds.append("(r.item_master_sku IS NULL OR r.item_master_sku='')"); 
    where = ("WHERE " + " AND ".join(conds)) if conds else ""

    total = db.execute(text(f"SELECT COUNT(*) FROM ps_rows r {where}"), params).scalar() or 0
    rows = db.execute(text(f"SELECT r.* FROM ps_rows r {where} ORDER BY r.id DESC LIMIT {per_page} OFFSET {(page-1)*per_page}"), params).fetchall()

    # KPIs
    kpi = db.execute(text(f"""SELECT
        SUM(CASE WHEN r.mapped_as='Sale' THEN r.quantity ELSE 0 END),
        SUM(CASE WHEN r.mapped_as='Sale' THEN r.invoice_amount ELSE 0 END),
        SUM(CASE WHEN r.mapped_as='Return' THEN r.quantity ELSE 0 END),
        SUM(CASE WHEN r.mapped_as='Return' THEN ABS(r.invoice_amount) ELSE 0 END)
        FROM ps_rows r {where}"""), params).fetchone()
    sq,sa,rq,ra = (kpi[0] or 0),(kpi[1] or 0),(kpi[2] or 0),(kpi[3] or 0)

    portals = [r[0] for r in db.execute(text("SELECT DISTINCT portal FROM ps_rows ORDER BY portal")).fetchall() if r[0]]

    return {
        "total": total, "page": page, "per_page": per_page,
        "portals": portals,
        "rows": [dict(r._mapping) for r in rows],
        "kpis": {"sale_qty":int(sq),"sale_amt":round(sa,2),"ret_qty":int(rq),"ret_amt":round(ra,2),
                 "net_qty":int(sq-rq),"net_amt":round(sa-ra,2)}
    }

@app.put("/api/portal-sale/row/{row_id}/mapped-as")
async def ps_update_row_mapped(row_id: int, request: Request,
                                db: Session = Depends(get_db),
                                current_user=Depends(auth.get_current_user)):
    from sqlalchemy import text
    d = await request.json()
    db.execute(text("UPDATE ps_rows SET mapped_as=:m WHERE id=:id"), {"m":d.get("mapped_as",""),"id":row_id})
    db.commit()
    return {"ok":True}

@app.post("/api/portal-sale/sku-map")
async def ps_new_sku_map(request: Request, db: Session = Depends(get_db),
                          current_user=Depends(auth.get_current_user)):
    from sqlalchemy import text
    d = await request.json()
    ps, po, im = d["portal_sku"], d["portal"], d["item_master_sku"]
    try:
        db.execute(text("INSERT INTO sku_mappings (portal_sku,portal,item_master_sku) VALUES (:ps,:po,:im) ON CONFLICT(portal_sku,portal) DO UPDATE SET item_master_sku=:im"), {"ps":ps,"po":po,"im":im})
    except:
        db.execute(text("UPDATE sku_mappings SET item_master_sku=:im WHERE portal_sku=:ps AND portal=:po"), {"im":im,"ps":ps,"po":po})
    db.execute(text("UPDATE ps_rows SET item_master_sku=:im WHERE portal_sku=:ps AND portal=:po"), {"im":im,"ps":ps,"po":po})
    db.commit()
    return {"ok":True}

@app.get("/api/portal-sale/export")
def ps_new_export(portal: str = None, mapped_as: str = None, token: str = None,
                   db: Session = Depends(get_db)):
    import csv, io
    from fastapi.responses import StreamingResponse
    from sqlalchemy import text
    from jose import jwt
    SECRET_KEY = "rajnandini-fashion-secret-key-2024"
    try:
        payload = jwt.decode(token, SECRET_KEY, algorithms=["HS256"])
        if not payload.get("sub"): raise HTTPException(401)
    except: raise HTTPException(401)
    ensure_portal_sale_tables(db)
    conds = []; params = {}
    if portal: conds.append("portal=:portal"); params["portal"] = portal
    if mapped_as: conds.append("mapped_as=:mapped_as"); params["mapped_as"] = mapped_as
    where = ("WHERE " + " AND ".join(conds)) if conds else ""
    rows = db.execute(text(f"SELECT * FROM ps_rows {where} ORDER BY id"), params).fetchall()
    output = io.StringIO()
    w = csv.writer(output)
    w.writerow(["Portal","File","Transaction Type","Mapped As","Portal SKU","Item Master SKU",
                "Order ID","Invoice No","Invoice Date","Qty","Amount","Ship From","Ship To","Fulfillment"])
    for r in rows:
        d = dict(r._mapping)
        w.writerow([d.get("portal",""),d.get("filename",""),d.get("transaction_type",""),
                    d.get("mapped_as",""),d.get("portal_sku",""),d.get("item_master_sku",""),
                    d.get("order_id",""),d.get("invoice_number",""),d.get("invoice_date",""),
                    d.get("quantity",0),d.get("invoice_amount",0),
                    d.get("ship_from_city",""),d.get("ship_to_city",""),d.get("fulfillment_channel","")])
    output.seek(0)
    from datetime import datetime
    fname = f"PortalSale_{datetime.now().strftime('%Y%m%d_%H%M')}.csv"
    return StreamingResponse(iter([output.getvalue()]),media_type="text/csv",
        headers={"Content-Disposition":f"attachment; filename={fname}"})

# ── ITEM MASTER ───────────────────────────────────────────────────────────────

SHEET_URL = "https://docs.google.com/spreadsheets/d/e/2PACX-1vR148FrPmuzy2vZ7EOa6NxpXgYiuWY92M-9aswGoZjI4QvDlkpnk6n9ajhG8LI8uLENniAAoWvrfaOR/pub?output=csv"

ITEM_COL_MAP = {
    "Category":         "category",
    "Uniware SKU":      "sku",
    "Style ID":         "style_id",
    "Availability":     "availability",
    "Uniware":          "uniware_stock",
    "FBA":              "fba_stock",
    "Sjit":             "sjit_stock",
    "FBF":              "fbf_stock",
    "Location":         "location",
    "Remark":           "remark",
    "ST CS":            "cost_price",
    "Wholesale Price":  "mrp",
    "Catalog Name":     "catalog_name",
}

@app.post("/api/sku-mappings/sync-from-portal")
def sync_sku_from_portal(db: Session = Depends(get_db), current_user=Depends(auth.get_current_user)):
    """Pull all unique portal SKUs from portal_sales into sku_mappings"""
    # Get all unique portal SKUs
    rows = db.execute(text("SELECT DISTINCT portal, sku FROM portal_sales WHERE sku IS NOT NULL AND sku != ''")).fetchall()
    # Get item master SKUs for direct match
    items = db.execute(text("SELECT sku FROM items")).fetchall()
    item_skus = {r.sku for r in items}
    added = 0
    mapped = 0
    for r in rows:
        portal_sku = r.sku
        portal = r.portal
        # Check if already in sku_mappings
        exists = db.execute(text("SELECT id, item_master_sku FROM sku_mappings WHERE portal_sku=:s AND portal=:p"),
                            {"s": portal_sku, "p": portal}).fetchone()
        if not exists:
            # Check direct match with item master
            im_sku = portal_sku if portal_sku in item_skus else None
            db.execute(text("INSERT INTO sku_mappings (portal_sku, portal, item_master_sku) VALUES (:s,:p,:u)"),
                       {"s": portal_sku, "p": portal, "u": im_sku})
            added += 1
            if im_sku: mapped += 1
        elif not exists.item_master_sku and portal_sku in item_skus:
            # Update if direct match found
            db.execute(text("UPDATE sku_mappings SET item_master_sku=:u WHERE portal_sku=:s AND portal=:p"),
                       {"u": portal_sku, "s": portal_sku, "p": portal})
            mapped += 1
    # Also update portal_sales with mapped SKUs
    mappings = db.execute(text("SELECT portal_sku, portal, item_master_sku FROM sku_mappings WHERE item_master_sku IS NOT NULL AND item_master_sku != ''")).fetchall()
    for m in mappings:
        db.execute(text("UPDATE portal_sales SET uniware_sku=:u WHERE sku=:s AND portal=:p AND (uniware_sku IS NULL OR uniware_sku='')"),
                   {"u": m.item_master_sku, "s": m.portal_sku, "p": m.portal})
    db.commit()
    return {"added": added, "auto_mapped": mapped, "total": len(rows)}

@app.delete("/api/portal-sales/delete/{file_id}")
def delete_portal_file(file_id: int, db: Session = Depends(get_db), current_user=Depends(auth.get_current_user)):
    pf = db.execute(text("SELECT * FROM portal_files WHERE id=:id"), {"id": file_id}).fetchone()
    if not pf: raise HTTPException(404, "File not found")
    pf = dict(pf._mapping)
    # Only delete portal_sales data — keep sku_mappings intact
    db.execute(text("DELETE FROM portal_sales WHERE file_name=:fn AND portal=:p"),
               {"fn": pf["file_name"], "p": pf["portal"]})
    db.execute(text("DELETE FROM portal_files WHERE id=:id"), {"id": file_id})
    db.commit()
    return {"success": True}

@app.get("/api/portal-sales/export")
def export_portal_sales(
    portal: str = None, mapped_type: str = None,
    month: str = None, year: str = None, token: str = None,
    db: Session = Depends(get_db)
):
    import csv as csvmod, io
    from fastapi.responses import StreamingResponse
    from jose import jwt, JWTError
    SECRET_KEY = "rajnandini-fashion-secret-key-2024"
    try:
        payload = jwt.decode(token, SECRET_KEY, algorithms=["HS256"])
        if not payload.get("sub"): raise HTTPException(401)
    except: raise HTTPException(401)
    q = "SELECT * FROM portal_sales WHERE 1=1"
    params = {}
    if portal and portal not in ("", "ALL"):
        q += " AND portal=:portal"; params["portal"] = portal
    if mapped_type and mapped_type not in ("", "ALL"):
        q += " AND mapped_type=:mt"; params["mt"] = mapped_type
    if month and month not in ("", "ALL"):
        q += " AND month=:month"; params["month"] = month.upper().strip()
    if year and str(year) not in ("", "ALL", "0"):
        q += " AND year=:year"; params["year"] = int(year)
    q += " ORDER BY invoice_date DESC"
    rows = db.execute(text(q), params).fetchall()
    output = io.StringIO()
    w = csvmod.writer(output)
    w.writerow(["Portal","Report Type","Transaction Type","Mapped As","SKU","Item Master SKU",
                "Order ID","Invoice No","Invoice Date","Qty","Amount","Tax Excl Gross",
                "Seller GSTIN","Fulfillment","Ship From","Ship To","Month","Year"])
    for r in rows:
        d = dict(r._mapping)
        w.writerow([d.get("portal",""), d.get("report_type",""), d.get("transaction_type",""),
                    d.get("mapped_type",""), d.get("sku",""), d.get("uniware_sku",""),
                    d.get("order_id",""), d.get("invoice_number",""), d.get("invoice_date",""),
                    d.get("quantity",0), d.get("invoice_amount",0), d.get("tax_exclusive_gross",0),
                    d.get("seller_gstin",""), d.get("fulfillment_channel",""),
                    d.get("ship_from_state",""), d.get("ship_to_state",""),
                    d.get("month",""), d.get("year","")])
    output.seek(0)
    from datetime import datetime
    fname = f"PortalSale_{datetime.now().strftime('%Y%m%d_%H%M')}.csv"
    return StreamingResponse(iter([output.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": f"attachment; filename={fname}"})





# ─── MYNTRA PORTAL SALE ──────────────────────────────────────────────────────

MYNTRA_ACCOUNTS = {
    "PPMP":   {"label": "Rajnandini MFN",  "fulfillment": "MFN", "portal": "Myntra"},
    "SJIT":   {"label": "Rajnandini SJIT", "fulfillment": "FBM", "portal": "Myntra"},
    "VIHAAN": {"label": "Vihaan",          "fulfillment": "MFN", "portal": "Myntra"},
}

def get_myntra_folder():
    username = os.environ.get("USERNAME") or os.environ.get("USER") or ""
    for base in [
        Path(f"C:/Users/{username}/OneDrive/Rajnandini/01 IMP_Do Not Edit  Delete/06 Portal Sales Files/Myntra"),
        Path(f"C:/Users/{username}/Rajnandini/01 IMP_Do Not Edit  Delete/06 Portal Sales Files/Myntra"),
    ]:
        if base.exists(): return base
    return Path(f"C:/Users/{username}/OneDrive/Rajnandini/01 IMP_Do Not Edit  Delete/06 Portal Sales Files/Myntra")

def detect_myntra_account(filename):
    fn = filename.upper()
    for acc in MYNTRA_ACCOUNTS:
        if acc in fn: return acc
    return "PPMP"

def detect_myntra_report_type(filename):
    fn = filename.upper().replace(' ', '_').replace('-', '_')
    if "RTO" in fn:    return "RTO"
    if "_RT." in fn or "_RT_" in fn or fn.endswith("_RT"): return "RT"
    if "B2C" in fn or "SALES_REVENUE" in fn or "REVENUE" in fn: return "B2C"
    if "PACKED" in fn: return "PACKED"
    return "UNKNOWN"

def parse_myntra_date(d):
    if not d or str(d).strip() in ('', '19700101'): return ""
    d = str(d).strip()
    try:
        from datetime import datetime as dt
        for fmt in ("%Y%m%d", "%Y-%m-%dT%H:%M:%S.%f%z", "%Y-%m-%dT%H:%M:%S%z", "%d/%m/%Y %H:%M"):
            try:
                parsed = dt.strptime(d[:19], fmt[:len(d[:19])])
                return parsed.strftime("%d-%m-%Y")
            except: pass
        if len(d) == 8 and d.isdigit():
            return f"{d[6:8]}-{d[4:6]}-{d[:4]}"
    except: pass
    return d

def parse_month_year_from_date(d):
    try:
        from datetime import datetime as dt
        d = str(d).strip()
        if len(d) == 8 and d.isdigit():
            dt_obj = dt.strptime(d, "%Y%m%d")
            return dt_obj.strftime("%B"), dt_obj.year
        for fmt in ("%Y-%m-%dT%H:%M:%S.%f%z", "%Y-%m-%dT%H:%M:%S%z"):
            try:
                dt_obj = dt.strptime(d[:25], fmt[:25])
                return dt_obj.strftime("%B"), dt_obj.year
            except: pass
    except: pass
    return None, None

@app.get("/api/myntra-sale/transaction-types")
def get_myntra_txn_types(db: Session = Depends(get_db), current_user=Depends(auth.get_current_user)):
    # Myntra uses report_type as transaction type: PACKED, RT, RTO
    rows = db.execute(text("SELECT DISTINCT report_type FROM myntra_sales WHERE report_type IS NOT NULL AND report_type!='B2C' ORDER BY report_type")).fetchall()
    types = [r[0] for r in rows]
    # Default mappings
    default_map = {"PACKED": "Sale", "RT": "Return", "RTO": "Return"}
    # Get saved mappings from transaction_type_mappings using portal='Myntra'
    mappings = db.execute(text("SELECT transaction_type, mapped_to FROM transaction_type_mappings WHERE portal='Myntra'")).fetchall()
    saved = {m[0]: m[1] for m in mappings}
    result = []
    for t in types:
        result.append({"report_type": t, "mapped_to": saved.get(t, default_map.get(t, "Sale"))})
    return result

@app.post("/api/myntra-sale/transaction-types/map")
def map_myntra_txn_type(data: dict, db: Session = Depends(get_db), current_user=Depends(auth.get_current_user)):
    from datetime import datetime
    rt = data["report_type"]; mapped = data["mapped_to"]
    exists = db.execute(text("SELECT id FROM transaction_type_mappings WHERE portal='Myntra' AND transaction_type=:t"), {"t": rt}).fetchone()
    if exists:
        db.execute(text("UPDATE transaction_type_mappings SET mapped_to=:m, updated_at=:u WHERE portal='Myntra' AND transaction_type=:t"),
                   {"m": mapped, "u": datetime.now(), "t": rt})
    else:
        db.execute(text("INSERT INTO transaction_type_mappings (portal, transaction_type, mapped_to, updated_at) VALUES ('Myntra',:t,:m,:u)"),
                   {"t": rt, "m": mapped, "u": datetime.now()})
    # Update mapped_type in myntra_sales
    db.execute(text("UPDATE myntra_sales SET mapped_type=:m WHERE report_type=:t"), {"m": mapped, "t": rt})
    db.commit()
    return {"success": True}

@app.get("/api/myntra-sale/scan")
def scan_myntra_folder(db: Session = Depends(get_db), current_user=Depends(auth.get_current_user)):
    from datetime import datetime
    folder = get_myntra_folder()
    if not folder.exists():
        return {"error": True, "folder_path": str(folder), "new": 0}
    new_files = []
    # Recursively find all CSV/XLSX files in all subfolders
    all_files = list(folder.rglob("*.csv")) + list(folder.rglob("*.xlsx"))
    for f in all_files:
        exists = db.execute(text("SELECT id FROM myntra_files WHERE file_path=:p"), {"p": str(f)}).fetchone()
        if not exists:
            account = detect_myntra_account(f.name)
            rtype   = detect_myntra_report_type(f.name)
            info    = MYNTRA_ACCOUNTS.get(account, {})
            # Extract month/period from parent folder name
            period  = f.parent.name if f.parent != folder else ""
            db.execute(text("""INSERT INTO myntra_files (account,file_name,file_path,report_type,fulfillment,period,status,row_count,detected_at)
                VALUES (:acc,:fn,:fp,:rt,:ful,:period,'pending',0,:dt)"""),
                {"acc": account, "fn": f.name, "fp": str(f), "rt": rtype,
                 "ful": info.get("fulfillment","MFN"), "period": period, "dt": datetime.now()})
            new_files.append(f.name)
    db.commit()
    rows = db.execute(text("SELECT * FROM myntra_files ORDER BY period DESC, account, report_type")).fetchall()
    return {"new": len(new_files), "files": [dict(r._mapping) for r in rows], "folder_path": str(folder)}

@app.get("/api/myntra-sale/files")
def get_myntra_files(db: Session = Depends(get_db), current_user=Depends(auth.get_current_user)):
    rows = db.execute(text("SELECT * FROM myntra_files ORDER BY detected_at DESC")).fetchall()
    return [dict(r._mapping) for r in rows]

@app.post("/api/myntra-sale/load/{file_id}")
def load_myntra_file(file_id: int, db: Session = Depends(get_db), current_user=Depends(auth.get_current_user)):
    import csv as csvmod
    from datetime import datetime
    pf = db.execute(text("SELECT * FROM myntra_files WHERE id=:id"), {"id": file_id}).fetchone()
    if not pf: raise HTTPException(404)
    pf = dict(pf._mapping)
    rt = pf["report_type"]
    if rt == "B2C":
        # B2C is reference only — index by packet_id for invoice lookup
        db.execute(text("DELETE FROM myntra_b2c WHERE file_name=:fn"), {"fn": pf["file_name"]})
        count = 0
        with open(pf["file_path"], encoding="utf-8-sig") as f:
            reader = list(csvmod.DictReader(f))
        for row in reader:
            if not any(v.strip() for v in row.values()): continue
            db.execute(text("""INSERT INTO myntra_b2c (file_name,account,packet_id,sale_order_code,invoice_number,sku_code,total_amount,taxable_amount,igst_amount,cgst_amount,sgst_amount,tcs_igst,tcs_cgst,tcs_sgst,tds_amount,order_created_date,packing_date,tracking_no,courier_code,customer_state,customer_pincode)
                VALUES (:fn,:acc,:pid,:soc,:inv,:sku,:ta,:taxa,:igst,:cgst,:sgst,:tcsi,:tcsc,:tcss,:tds,:ocd,:pd,:tn,:cc,:cs,:cp)"""),
                {"fn": pf["file_name"], "acc": pf["account"],
                 "pid": row.get("packed_id","").strip(), "soc": row.get("Sale_Order_Code","").strip(),
                 "inv": row.get("Invoice_Number","").strip(), "sku": row.get("SKU_Code","").strip(),
                 "ta": row.get("Total_Amount","0") or 0, "taxa": row.get("Taxable_Amount","0") or 0,
                 "igst": row.get("Igst_Amount","0") or 0, "cgst": row.get("Cgst_Amount","0") or 0,
                 "sgst": row.get("Sgst_Amount","0") or 0,
                 "tcsi": row.get("Tcs_Igst_Amount","0") or 0, "tcsc": row.get("Tcs_Cgst_Amount","0") or 0,
                 "tcss": row.get("Tcs_Sgst_Amount","0") or 0, "tds": row.get("tds_amount","0") or 0,
                 "ocd": row.get("Order_Created_Date","").strip()[:10],
                 "pd": row.get("Packing_Date","").strip(), "tn": row.get("Tracking_no","").strip(),
                 "cc": row.get("Courier_Code","").strip(), "cs": row.get("Customer_State","").strip(),
                 "cp": row.get("Customer_PinCode","").strip()})
            count += 1
        db.execute(text("UPDATE myntra_files SET status='loaded',row_count=:c,loaded_at=:t WHERE id=:id"),
                   {"c": count, "t": datetime.now(), "id": file_id})
        db.commit()
        return {"success": True, "rows": count, "type": "B2C reference"}

    # PACKED / RT / RTO
    db.execute(text("DELETE FROM myntra_sales WHERE file_name=:fn"), {"fn": pf["file_name"]})
    count = 0
    with open(pf["file_path"], encoding="utf-8-sig") as f:
        reader = list(csvmod.DictReader(f))
    # Build B2C invoice lookup (packet_id → invoice_number, sku_code)
    b2c_lookup = {}
    b2c_rows = db.execute(text("SELECT packet_id, invoice_number, sku_code FROM myntra_b2c WHERE account=:acc"),
                          {"acc": pf["account"]}).fetchall()
    for b in b2c_rows: b2c_lookup[b[0]] = {"invoice": b[1], "sku": b[2]}

    mapped_type = {"PACKED": "Sale", "RT": "Return", "RTO": "Return", "B2C": "Sale"}.get(rt, "Sale")
    fulfillment = pf["fulfillment"]
    account     = pf["account"]
    acc_info    = MYNTRA_ACCOUNTS.get(account, {})

    for row in reader:
        if not any(str(v).strip() for v in row.values()): continue
        # Get packet_id and order_id — RT uses old_parent_order_id, RTO uses order_id
        packet_id = str(row.get("packet_id") or row.get("shipment_id") or "").strip()
        order_id  = str(row.get("order_id") or row.get("old_parent_order_id") or row.get("shipment_id") or "").strip()

        # Date
        if rt == "PACKED":
            raw_date = row.get("order_packed_date","") or row.get("order_shipped_date","")
        elif rt == "RT":
            raw_date = row.get("fr_refunded_date","") or row.get("order_delivered_date","")
        else:
            raw_date = row.get("order_cancel_date","") or row.get("order_packed_date","")
        invoice_date = parse_myntra_date(raw_date)
        month, year = parse_month_year_from_date(raw_date)

        # SKU — from B2C if available
        b2c = b2c_lookup.get(packet_id, {})
        sku  = b2c.get("sku","") or str(row.get("sku_id","")).strip()
        inv  = b2c.get("invoice","")

        qty_raw = row.get("quantity") or row.get("Qty") or "1"
        try: qty = int(float(str(qty_raw).strip() or 1))
        except: qty = 1

        # Amounts
        def gf(k): 
            v = str(row.get(k,"") or "0").strip()
            try: return float(v) if v else 0.0
            except: return 0.0

        shipment_value = gf("shipment_value") or gf("invoiceamount") or gf("final_amount")
        seller_price   = gf("seller_price")
        base_value     = gf("base_value")
        tax_amount     = gf("tax_amount")
        gta_fee        = gf("gta_fee")
        tds_amount     = gf("tds_amount")
        tcs_amount     = gf("tcs_amount")
        igst_amt       = gf("igst_amt") or gf("igst_amount")
        cgst_amt       = gf("cgst_amt") or gf("cgst_amount")
        sgst_amt       = gf("sgst_amt") or gf("sgst_amount")
        mrp            = gf("mrp")
        discount       = gf("discount")

        # For returns, negate amounts
        sign = -1 if mapped_type == "Return" else 1

        db.execute(text("""INSERT INTO myntra_sales
            (portal, account, fulfillment, file_name, report_type, mapped_type,
             order_id, packet_id, sku, invoice_number, invoice_date, month, year, quantity,
             shipment_value, seller_price, base_value, tax_amount, gta_fee, tds_amount, tcs_amount,
             igst_amt, cgst_amt, sgst_amt, mrp, discount,
             seller_gstin, warehouse_name, ship_to_state, payment_method, brand, article_type)
            VALUES (:portal,:acc,:ful,:fn,:rt,:mt,
             :oid,:pid,:sku,:inv,:idate,:month,:year,:qty,
             :sv,:sp,:bv,:ta,:gta,:tds,:tcs,
             :igst,:cgst,:sgst,:mrp,:disc,
             :gstin,:wh,:ship_to,:pay,:brand,:art)"""),
            {"portal": "Myntra", "acc": account, "ful": fulfillment,
             "fn": pf["file_name"], "rt": rt, "mt": mapped_type,
             "oid": order_id, "pid": packet_id, "sku": sku, "inv": inv,
             "idate": invoice_date, "month": month, "year": year, "qty": qty,
             "sv": round(shipment_value*sign,2), "sp": seller_price, "bv": base_value,
             "ta": round(tax_amount*sign,2), "gta": gta_fee, "tds": tds_amount, "tcs": tcs_amount,
             "igst": round(igst_amt*sign,2), "cgst": round(cgst_amt*sign,2), "sgst": round(sgst_amt*sign,2),
             "mrp": mrp, "disc": discount,
             "gstin": row.get("seller_gstin","").strip(),
             "wh": row.get("warehouse_name","").strip(),
             "ship_to": (row.get("customer_delivery_state_code","") or row.get("location","") or row.get("delivery_state","") or row.get("customer_state","")).strip(),
             "pay": row.get("payment_method","").strip(),
             "brand": row.get("brand","").strip(), "art": row.get("article_type","").strip()})
        count += 1

    db.execute(text("UPDATE myntra_files SET status='loaded',row_count=:c,loaded_at=:t WHERE id=:id"),
               {"c": count, "t": datetime.now(), "id": file_id})
    # Auto-apply any saved transaction type mappings
    mappings = db.execute(text(
        "SELECT transaction_type, mapped_to FROM transaction_type_mappings WHERE portal='Myntra'"
    )).fetchall()
    for m in mappings:
        db.execute(text("UPDATE myntra_sales SET mapped_type=:mt WHERE report_type=:rt AND file_name=:fn"),
                   {"mt": m.mapped_to, "rt": m.transaction_type, "fn": pf["file_name"]})
    db.commit()
    return {"success": True, "rows": count, "type": rt}

@app.post("/api/myntra-sale/apply-mapping")
def myntra_apply_mapping(db: Session = Depends(get_db), current_user=Depends(auth.get_current_user)):
    """Apply transaction_type_mappings to myntra_sales.mapped_type"""
    mappings = db.execute(text(
        "SELECT transaction_type, mapped_to FROM transaction_type_mappings WHERE portal='Myntra'"
    )).fetchall()
    updated = 0
    for m in mappings:
        r = db.execute(text("""UPDATE myntra_sales SET mapped_type=:mt WHERE report_type=:rt"""),
                       {"mt": m.mapped_to, "rt": m.transaction_type})
        updated += r.rowcount
    db.commit()
    return {"updated": updated}

@app.delete("/api/myntra-sale/delete/{file_id}")
def delete_myntra_file(file_id: int, db: Session = Depends(get_db), current_user=Depends(auth.get_current_user)):
    pf = db.execute(text("SELECT * FROM myntra_files WHERE id=:id"), {"id": file_id}).fetchone()
    if not pf: raise HTTPException(404)
    pf = dict(pf._mapping)
    if pf["report_type"] == "B2C":
        db.execute(text("DELETE FROM myntra_b2c WHERE file_name=:fn"), {"fn": pf["file_name"]})
    else:
        db.execute(text("DELETE FROM myntra_sales WHERE file_name=:fn"), {"fn": pf["file_name"]})
    db.execute(text("DELETE FROM myntra_files WHERE id=:id"), {"id": file_id})
    db.commit()
    return {"success": True}

@app.get("/api/myntra-sale/summary")
def myntra_summary(db: Session = Depends(get_db), current_user=Depends(auth.get_current_user)):
    row = db.execute(text("""SELECT COUNT(*),
        SUM(CASE WHEN mapped_type='Sale' THEN quantity ELSE 0 END),
        SUM(CASE WHEN mapped_type='Return' THEN ABS(quantity) ELSE 0 END),
        SUM(CASE WHEN mapped_type='Sale' THEN shipment_value ELSE 0 END),
        SUM(CASE WHEN mapped_type='Return' THEN ABS(shipment_value) ELSE 0 END)
        FROM myntra_sales""")).fetchone()
    unmapped = db.execute(text("SELECT COUNT(DISTINCT sku) FROM myntra_sales WHERE sku IS NOT NULL AND sku!='' AND sku NOT IN (SELECT sku FROM items) AND sku NOT IN (SELECT portal_sku FROM sku_mappings WHERE portal='Myntra')")).scalar() or 0
    return {
        "total": row[0] or 0, "sale_qty": row[1] or 0, "return_qty": row[2] or 0,
        "sale_amt": row[3] or 0, "return_amt": row[4] or 0,
        "net_qty": (row[1] or 0) - (row[2] or 0),
        "net_amt": (row[3] or 0) - (row[4] or 0),
        "unmapped_skus": unmapped
    }

@app.get("/api/myntra-sale/data")
def myntra_data(account: str = None, report_type: str = None, month: str = None, year: str = None,
                mapped_type: str = None, search: str = None, page: int = 1, per_page: int = 25,
                db: Session = Depends(get_db), current_user=Depends(auth.get_current_user)):
    q = "SELECT * FROM myntra_sales WHERE 1=1"
    p = {}
    if account:     q += " AND account=:acc";           p["acc"] = account
    if report_type: q += " AND report_type=:rt";        p["rt"]  = report_type
    if month:       q += " AND month=:month";           p["month"] = month
    if year and str(year).strip(): q += " AND year=:year"; p["year"] = int(year)
    if mapped_type: q += " AND mapped_type=:mt";        p["mt"]  = mapped_type
    if search:
        q += " AND (sku LIKE :s OR order_id LIKE :s OR invoice_number LIKE :s OR packet_id LIKE :s)"
        p["s"] = f"%{search}%"
    total = db.execute(text(q.replace("SELECT *","SELECT COUNT(*)")), p).scalar()
    q += f" ORDER BY invoice_date DESC LIMIT {per_page} OFFSET {(page-1)*per_page}"
    rows = db.execute(text(q), p).fetchall()
    return {"total": total, "page": page, "per_page": per_page, "items": [dict(r._mapping) for r in rows]}

@app.get("/api/myntra-sale/export")
def myntra_export(account: str = None, report_type: str = None, month: str = None, year: str = None,
                  token: str = None, db: Session = Depends(get_db)):
    import csv as csvmod, io
    from fastapi.responses import StreamingResponse
    from jose import jwt
    SECRET_KEY = "rajnandini-fashion-secret-key-2024"
    try:
        payload = jwt.decode(token, SECRET_KEY, algorithms=["HS256"])
        if not payload.get("sub"): raise HTTPException(401)
    except: raise HTTPException(401)
    q = "SELECT * FROM myntra_sales WHERE 1=1"
    p = {}
    if account:     q += " AND account=:acc";    p["acc"]   = account
    if report_type: q += " AND report_type=:rt"; p["rt"]    = report_type
    if month:       q += " AND month=:month";    p["month"] = month
    if year and str(year).strip(): q += " AND year=:year"; p["year"] = int(year)
    q += " ORDER BY invoice_date DESC"
    rows = db.execute(text(q), p).fetchall()
    output = io.StringIO()
    w = csvmod.writer(output)
    if rows:
        w.writerow(rows[0]._mapping.keys())
        for r in rows: w.writerow(list(r._mapping.values()))
    output.seek(0)
    from datetime import datetime
    return StreamingResponse(iter([output.getvalue()]), media_type="text/csv",
        headers={"Content-Disposition": f"attachment; filename=Myntra_Sale_{datetime.now().strftime('%Y%m%d')}.csv"})

# ─── PORTAL EXTRA DATA ───────────────────────────────────────────────────────

EXTRA_DATA_REPORT_TYPES = {
    'fba-returns':      'FBA Customer Returns',
    'storage-fees':     'Monthly Storage Fees',
    'order-reimb':      'Order Level Reimbursement',
    'reimbursements':   'Reimbursements',
    'replacements':     'Replacements',
    'aged-inv':         'Aged Inventory Surcharge',
}

def get_extra_data_folder():
    username = os.environ.get("USERNAME") or os.environ.get("USER") or ""
    candidates = [
        Path(f"C:/Users/{username}/OneDrive/Rajnandini/01 IMP_Do Not Edit  Delete/08 Portal Extra Data"),
        Path(f"C:/Users/{username}/Rajnandini/01 IMP_Do Not Edit  Delete/08 Portal Extra Data"),
    ]
    for p in candidates:
        if p.exists(): return p
    return candidates[0]

def detect_report_type(filename, folder_name):
    """Detect report type from folder name"""
    fn = folder_name.lower()
    if 'fba customer' in fn or 'fba_customer' in fn: return 'fba-returns'
    if 'storage' in fn: return 'storage-fees'
    if 'order level' in fn or 'order_level' in fn: return 'order-reimb'
    if 'reimbursement' in fn and 'order' not in fn: return 'reimbursements'
    if 'replacement' in fn: return 'replacements'
    if 'aged' in fn: return 'aged-inv'
    return 'unknown'

@app.get("/api/extra-data/scan")
def scan_extra_data(db: Session = Depends(get_db), current_user=Depends(auth.get_current_user)):
    from datetime import datetime
    folder = get_extra_data_folder()
    if not folder.exists():
        return {"error": True, "folder_path": str(folder), "new": 0}
    new_files = []
    for portal_dir in folder.iterdir():
        if not portal_dir.is_dir(): continue
        portal = portal_dir.name
        for sub_dir in portal_dir.iterdir():
            if not sub_dir.is_dir(): continue
            report_type = detect_report_type("", sub_dir.name)
            for f in list(sub_dir.glob("*.csv")) + list(sub_dir.glob("*.xlsx")):
                exists = db.execute(text("SELECT id FROM extra_data_files WHERE file_path=:p"), {"p": str(f)}).fetchone()
                if not exists:
                    db.execute(text("""INSERT INTO extra_data_files (portal,file_name,file_path,report_type,status,row_count,detected_at)
                        VALUES (:portal,:fn,:fp,:rt,'pending',0,:dt)"""),
                        {"portal":portal,"fn":f.name,"fp":str(f),"rt":report_type,"dt":datetime.now()})
                    new_files.append(f.name)
    db.commit()
    rows = db.execute(text("SELECT * FROM extra_data_files ORDER BY detected_at DESC")).fetchall()
    result = []
    for r in rows:
        d = dict(r._mapping)
        d['report_type_label'] = EXTRA_DATA_REPORT_TYPES.get(d['report_type'], d['report_type'])
        result.append(d)
    return {"new": len(new_files), "files": result, "folder_path": str(folder)}

@app.get("/api/extra-data/files")
def get_extra_data_files(db: Session = Depends(get_db), current_user=Depends(auth.get_current_user)):
    rows = db.execute(text("SELECT * FROM extra_data_files ORDER BY detected_at DESC")).fetchall()
    result = []
    for r in rows:
        d = dict(r._mapping)
        d['report_type_label'] = EXTRA_DATA_REPORT_TYPES.get(d['report_type'], d['report_type'])
        result.append(d)
    return result

@app.post("/api/extra-data/load/{file_id}")
def load_extra_data_file(file_id: int, db: Session = Depends(get_db), current_user=Depends(auth.get_current_user)):
    import csv as csvmod
    from datetime import datetime
    pf = db.execute(text("SELECT * FROM extra_data_files WHERE id=:id"), {"id": file_id}).fetchone()
    if not pf: raise HTTPException(404)
    pf = dict(pf._mapping)
    rt = pf['report_type']
    table_map = {
        'fba-returns': 'ed_fba_returns',
        'storage-fees': 'ed_storage_fees',
        'order-reimb': 'ed_order_reimb',
        'reimbursements': 'ed_reimbursements',
        'replacements': 'ed_replacements',
        'aged-inv': 'ed_aged_inv',
    }
    table = table_map.get(rt)
    if not table: raise HTTPException(400, f"Unknown report type: {rt}")
    db.execute(text(f"DELETE FROM {table} WHERE file_name=:fn AND portal=:p"),
               {"fn": pf["file_name"], "p": pf["portal"]})
    count = 0
    try:
        fname = pf["file_path"]
        with open(fname, encoding="utf-8-sig") as f:
            reader = list(csvmod.reader(f))
        headers = [h.strip().lower().replace(' ','_').replace('/','_').replace('-','_').replace('(','').replace(')','') for h in reader[0]]
        col_map = {h: i for i, h in enumerate(headers)}
        def g(row, col):
            idx = col_map.get(col.lower().replace(' ','_').replace('/','_').replace('-','_').replace('(','').replace(')',''))
            return row[idx].strip() if idx is not None and idx < len(row) else ""
        def gf(row, col):
            try: return float(g(row, col)) if g(row, col) else 0.0
            except: return 0.0
        for row in reader[1:]:
            if not any(c.strip() for c in row): continue
            if rt == 'fba-returns':
                db.execute(text("""INSERT INTO ed_fba_returns (portal,file_name,return_date,order_id,sku,asin,fnsku,product_name,quantity,fulfillment_center_id,detailed_disposition,reason,customer_comments)
                    VALUES (:portal,:fn,:rd,:oid,:sku,:asin,:fnsku,:pn,:qty,:fc,:disp,:reason,:cc)"""),
                    {"portal":pf["portal"],"fn":pf["file_name"],"rd":g(row,"return-date"),"oid":g(row,"order-id"),
                     "sku":g(row,"sku"),"asin":g(row,"asin"),"fnsku":g(row,"fnsku"),"pn":g(row,"product-name"),
                     "qty":g(row,"quantity"),"fc":g(row,"fulfillment-center-id"),"disp":g(row,"detailed-disposition"),
                     "reason":g(row,"reason"),"cc":g(row,"customer-comments")})
            elif rt == 'storage-fees':
                db.execute(text("""INSERT INTO ed_storage_fees (portal,file_name,asin,fnsku,product_name,fulfillment_center,month_of_charge,average_quantity_on_hand,item_volume,storage_rate,estimated_monthly_storage_fee)
                    VALUES (:portal,:fn,:asin,:fnsku,:pn,:fc,:month,:avgqty,:vol,:rate,:fee)"""),
                    {"portal":pf["portal"],"fn":pf["file_name"],"asin":g(row,"asin"),"fnsku":g(row,"fnsku"),
                     "pn":g(row,"product-name"),"fc":g(row,"fulfillment-center"),"month":g(row,"month-of-charge"),
                     "avgqty":gf(row,"average-quantity-on-hand"),"vol":gf(row,"item-volume"),
                     "rate":gf(row,"storage-rate"),"fee":gf(row,"estimated-monthly-storage-fee")})
            elif rt == 'order-reimb':
                db.execute(text("""INSERT INTO ed_order_reimb (portal,file_name,order_id,asin,product_title,msku,fnsku,shipped_quantity,refund_date,customer_return_reason,return_item_condition,reimbursement_reason,reimbursement_status,reimbursed_amount_per_unit)
                    VALUES (:portal,:fn,:oid,:asin,:pt,:msku,:fnsku,:qty,:dt,:crr,:ric,:rr,:rs,:amt)"""),
                    {"portal":pf["portal"],"fn":pf["file_name"],"oid":g(row,"Order_ID"),"asin":g(row,"ASIN"),
                     "pt":g(row,"Product_Title"),"msku":g(row,"MSKU"),"fnsku":g(row,"FNSKU"),
                     "qty":g(row,"Shipped_Quantity"),"dt":g(row,"Refund_Replacement_Date"),
                     "crr":g(row,"Customer_Return_Reason"),"ric":g(row,"Return_Item_Condition"),
                     "rr":g(row,"Reimbursement_Reason"),"rs":g(row,"Reimbursement_Status"),
                     "amt":gf(row,"Reimbursed_Amount_Per_Unit")})
            elif rt == 'reimbursements':
                db.execute(text("""INSERT INTO ed_reimbursements (portal,file_name,approval_date,reimbursement_id,case_id,amazon_order_id,reason,sku,fnsku,asin,product_name,condition,currency,amount_per_unit,amount_total,quantity_reimbursed_cash,quantity_reimbursed_inventory,quantity_reimbursed_total)
                    VALUES (:portal,:fn,:dt,:rid,:cid,:oid,:reason,:sku,:fnsku,:asin,:pn,:cond,:cur,:apu,:at,:qrc,:qri,:qrt)"""),
                    {"portal":pf["portal"],"fn":pf["file_name"],"dt":g(row,"approval-date"),"rid":g(row,"reimbursement-id"),
                     "cid":g(row,"case-id"),"oid":g(row,"amazon-order-id"),"reason":g(row,"reason"),
                     "sku":g(row,"sku"),"fnsku":g(row,"fnsku"),"asin":g(row,"asin"),"pn":g(row,"product-name"),
                     "cond":g(row,"condition"),"cur":g(row,"currency-unit"),"apu":gf(row,"amount-per-unit"),
                     "at":gf(row,"amount-total"),"qrc":gf(row,"quantity-reimbursed-cash"),
                     "qri":gf(row,"quantity-reimbursed-inventory"),"qrt":gf(row,"quantity-reimbursed-total")})
            elif rt == 'replacements':
                db.execute(text("""INSERT INTO ed_replacements (portal,file_name,shipment_date,sku,asin,fulfillment_center_id,original_fulfillment_center_id,quantity,replacement_reason_code,replacement_amazon_order_id,original_amazon_order_id)
                    VALUES (:portal,:fn,:dt,:sku,:asin,:fc,:ofc,:qty,:rc,:raid,:oaid)"""),
                    {"portal":pf["portal"],"fn":pf["file_name"],"dt":g(row,"shipment-date"),"sku":g(row,"sku"),
                     "asin":g(row,"asin"),"fc":g(row,"fulfillment-center-id"),"ofc":g(row,"original-fulfillment-center-id"),
                     "qty":gf(row,"quantity"),"rc":g(row,"replacement-reason-code"),
                     "raid":g(row,"replacement-amazon-order-id"),"oaid":g(row,"original-amazon-order-id")})
            elif rt == 'aged-inv':
                db.execute(text("""INSERT INTO ed_aged_inv (portal,file_name,asin,sku,product_name,fulfillment_center,quantity,age_days,surcharge_per_unit,total_surcharge)
                    VALUES (:portal,:fn,:asin,:sku,:pn,:fc,:qty,:age,:spu,:ts)"""),
                    {"portal":pf["portal"],"fn":pf["file_name"],"asin":g(row,"asin"),"sku":g(row,"sku"),
                     "pn":g(row,"product-name"),"fc":g(row,"fulfillment-center-id"),
                     "qty":gf(row,"quantity"),"age":gf(row,"weeks-of-cover-t0"),
                     "spu":gf(row,"surcharge-per-unit"),"ts":gf(row,"total-surcharge")})
            count += 1
        db.execute(text("UPDATE extra_data_files SET status='loaded',row_count=:c,loaded_at=:t WHERE id=:id"),
                   {"c":count,"t":datetime.now(),"id":file_id})
        db.commit()
        return {"success": True, "rows": count}
    except Exception as e:
        db.rollback()
        raise HTTPException(500, str(e))

@app.delete("/api/extra-data/delete/{file_id}")
def delete_extra_data_file(file_id: int, db: Session = Depends(get_db), current_user=Depends(auth.get_current_user)):
    pf = db.execute(text("SELECT * FROM extra_data_files WHERE id=:id"), {"id": file_id}).fetchone()
    if not pf: raise HTTPException(404)
    pf = dict(pf._mapping)
    table_map = {'fba-returns':'ed_fba_returns','storage-fees':'ed_storage_fees','order-reimb':'ed_order_reimb',
                 'reimbursements':'ed_reimbursements','replacements':'ed_replacements','aged-inv':'ed_aged_inv'}
    table = table_map.get(pf['report_type'])
    if table:
        db.execute(text(f"DELETE FROM {table} WHERE file_name=:fn AND portal=:p"), {"fn":pf["file_name"],"p":pf["portal"]})
    db.execute(text("DELETE FROM extra_data_files WHERE id=:id"), {"id": file_id})
    db.commit()
    return {"success": True}

@app.get("/api/extra-data/data/{report_type}")
def get_extra_data(report_type: str, search: str = None, page: int = 1, per_page: int = 50,
                   db: Session = Depends(get_db), current_user=Depends(auth.get_current_user)):
    table_map = {'fba-returns':'ed_fba_returns','storage-fees':'ed_storage_fees','order-reimb':'ed_order_reimb',
                 'reimbursements':'ed_reimbursements','replacements':'ed_replacements','aged-inv':'ed_aged_inv'}
    search_cols = {
        'fba-returns': ['sku','order_id','reason','detailed_disposition'],
        'storage-fees': ['asin','fnsku','product_name'],
        'order-reimb': ['msku','order_id','asin','reimbursement_status'],
        'reimbursements': ['sku','amazon_order_id','reason'],
        'replacements': ['sku','asin','replacement_amazon_order_id','original_amazon_order_id'],
        'aged-inv': ['asin','sku','product_name'],
    }
    table = table_map.get(report_type)
    if not table: raise HTTPException(400)
    q = f"SELECT * FROM {table} WHERE 1=1"
    params = {}
    if search:
        cols = search_cols.get(report_type, ['sku'])
        q += " AND (" + " OR ".join([f"{c} LIKE :s" for c in cols]) + ")"
        params["s"] = f"%{search}%"
    total = db.execute(text(q.replace("SELECT *","SELECT COUNT(*)")), params).scalar()
    # totals
    extra = {}
    if report_type == 'storage-fees':
        r = db.execute(text(f"SELECT SUM(estimated_monthly_storage_fee) FROM {table} WHERE 1=1" + (" AND product_name LIKE :s" if search else "")), params).scalar()
        extra['total_fee'] = r or 0
    if report_type == 'reimbursements':
        r = db.execute(text(f"SELECT SUM(amount_total) FROM {table} WHERE 1=1" + (" AND sku LIKE :s" if search else "")), params).scalar()
        extra['total_amount'] = r or 0
    q += f" ORDER BY id DESC LIMIT {per_page} OFFSET {(page-1)*per_page}"
    rows = db.execute(text(q), params).fetchall()
    return {"total": total, "page": page, "per_page": per_page,
            "items": [dict(r._mapping) for r in rows], **extra}

@app.get("/api/extra-data/export/{report_type}")
def export_extra_data(report_type: str, token: str = None, db: Session = Depends(get_db)):
    import csv as csvmod, io
    from fastapi.responses import StreamingResponse
    from jose import jwt
    SECRET_KEY = "rajnandini-fashion-secret-key-2024"
    try:
        payload = jwt.decode(token, SECRET_KEY, algorithms=["HS256"])
        if not payload.get("sub"): raise HTTPException(401)
    except: raise HTTPException(401)
    table_map = {'fba-returns':'ed_fba_returns','storage-fees':'ed_storage_fees','order-reimb':'ed_order_reimb',
                 'reimbursements':'ed_reimbursements','replacements':'ed_replacements','aged-inv':'ed_aged_inv'}
    table = table_map.get(report_type)
    if not table: raise HTTPException(400)
    rows = db.execute(text(f"SELECT * FROM {table} ORDER BY id DESC")).fetchall()
    output = io.StringIO()
    w = csvmod.writer(output)
    if rows:
        w.writerow(rows[0]._mapping.keys())
        for r in rows: w.writerow(list(r._mapping.values()))
    output.seek(0)
    from datetime import datetime
    fname = f"{report_type}_{datetime.now().strftime('%Y%m%d_%H%M')}.csv"
    return StreamingResponse(iter([output.getvalue()]), media_type="text/csv",
        headers={"Content-Disposition": f"attachment; filename={fname}"})

# ─── PORTAL PAYMENT ──────────────────────────────────────────────────────────

def get_portal_payment_folder():
    username = os.environ.get("USERNAME") or os.environ.get("USER") or ""
    candidates = [
        Path(f"C:/Users/{username}/OneDrive/Rajnandini/01 IMP_Do Not Edit  Delete/07 Portal Payment File"),
        Path(f"C:/Users/{username}/Rajnandini/01 IMP_Do Not Edit  Delete/07 Portal Payment File"),
    ]
    for p in candidates:
        if p.exists():
            return p
    return candidates[0]

def parse_payment_row(row, headers, portal, file_name):
    def g(col):
        for i, h in enumerate(headers):
            if h.strip().lower() == col.lower():
                v = row[i] if i < len(row) else ""
                return str(v).strip() if v is not None else ""
        return ""
    def gf(col):
        v = g(col)
        try: return float(v) if v else 0.0
        except: return 0.0
    # Parse date
    raw_date = g("date/time")
    try:
        from datetime import datetime as dt
        for fmt in ("%d %b %Y %I:%M:%S %p UTC", "%Y-%m-%d %H:%M:%S", "%d/%m/%Y"):
            try:
                d = dt.strptime(raw_date.strip(), fmt)
                raw_date = d.strftime("%d-%m-%Y")
                month = d.strftime("%B")
                year = d.year
                break
            except: month = None; year = None
    except: month = None; year = None
    return dict(
        portal=portal, file_name=file_name,
        date_time=raw_date, month=month, year=year,
        settlement_id=g("settlement id"),
        type=g("type"), order_id=g("order id"), sku=g("Sku"),
        description=g("description"), quantity=g("quantity"),
        marketplace=g("marketplace"), fulfillment=g("fulfillment"),
        order_city=g("order city"), order_state=g("order state"),
        product_sales=gf("product sales"),
        shipping_credits=gf("shipping credits"),
        promotional_rebates=gf("promotional rebates"),
        selling_fees=gf("selling fees"),
        fba_fees=gf("fba fees"),
        other_transaction_fees=gf("other transaction fees"),
        other=gf("other"),
        total=gf("total"),
    )

@app.get("/api/portal-payment/scan")
def scan_payment_folder(db: Session = Depends(get_db), current_user=Depends(auth.get_current_user)):
    folder = get_portal_payment_folder()
    if not folder.exists():
        return {"error": True, "folder_path": str(folder), "files": [], "new": 0}
    new_files = []
    for portal_dir in folder.iterdir():
        if not portal_dir.is_dir(): continue
        portal = portal_dir.name
        for f in list(portal_dir.glob("*.csv")) + list(portal_dir.glob("*.xlsx")):
            exists = db.execute(text("SELECT id FROM payment_files WHERE file_path=:p"), {"p": str(f)}).fetchone()
            if not exists:
                from datetime import datetime
                db.execute(text("""INSERT INTO payment_files (portal,file_name,file_path,status,row_count,detected_at)
                    VALUES (:portal,:fn,:fp,'pending',0,:dt)"""),
                    {"portal":portal,"fn":f.name,"fp":str(f),"dt":datetime.now()})
                new_files.append(f.name)
    db.commit()
    rows = db.execute(text("SELECT * FROM payment_files ORDER BY detected_at DESC")).fetchall()
    return {"new": len(new_files), "files": [dict(r._mapping) for r in rows], "folder_path": str(folder)}

@app.get("/api/portal-payment/files")
def get_payment_files(db: Session = Depends(get_db), current_user=Depends(auth.get_current_user)):
    rows = db.execute(text("SELECT * FROM payment_files ORDER BY detected_at DESC")).fetchall()
    return [dict(r._mapping) for r in rows]

@app.post("/api/portal-payment/load/{file_id}")
def load_payment_file(file_id: int, db: Session = Depends(get_db), current_user=Depends(auth.get_current_user)):
    import csv as csvmod
    from datetime import datetime
    pf = db.execute(text("SELECT * FROM payment_files WHERE id=:id"), {"id": file_id}).fetchone()
    if not pf: raise HTTPException(404, "File not found")
    pf = dict(pf._mapping)
    db.execute(text("DELETE FROM payment_transactions WHERE file_name=:fn AND portal=:p"),
               {"fn": pf["file_name"], "p": pf["portal"]})
    count = 0
    try:
        fname = pf["file_path"]
        # Find header row — Amazon payment files have preamble rows
        if fname.endswith(".xlsx"):
            import openpyxl
            wb = openpyxl.load_workbook(fname, read_only=True, data_only=True)
            ws = wb.active
            all_rows = list(ws.iter_rows(values_only=True))
            # Find header row
            header_idx = 0
            for i, row in enumerate(all_rows):
                if row and any(str(c or '').lower() in ['date/time','type','settlement id'] for c in row):
                    header_idx = i; break
            headers = [str(h or '').strip() for h in all_rows[header_idx]]
            data_rows = [[str(c or '').strip() for c in r] for r in all_rows[header_idx+1:]]
        else:
            with open(fname, encoding="utf-8-sig") as f:
                all_lines = f.readlines()
            header_idx = 0
            for i, line in enumerate(all_lines):
                if 'date/time' in line.lower() or 'settlement id' in line.lower():
                    header_idx = i; break
            content = ''.join(all_lines[header_idx:])
            reader = list(csvmod.reader(content.splitlines()))
            headers = reader[0]
            data_rows = reader[1:]
        for row in data_rows:
            if not any(str(c).strip() for c in row): continue
            data = parse_payment_row(row, headers, pf["portal"], pf["file_name"])
            cols = ','.join(data.keys())
            vals = ','.join([f':{k}' for k in data.keys()])
            db.execute(text(f"INSERT INTO payment_transactions ({cols}) VALUES ({vals})"), data)
            count += 1
        db.execute(text("UPDATE payment_files SET status='loaded',row_count=:c,loaded_at=:t WHERE id=:id"),
                   {"c": count, "t": datetime.now(), "id": file_id})
        db.commit()
        return {"success": True, "rows": count}
    except Exception as e:
        db.rollback()
        raise HTTPException(500, str(e))

@app.delete("/api/portal-payment/delete/{file_id}")
def delete_payment_file(file_id: int, db: Session = Depends(get_db), current_user=Depends(auth.get_current_user)):
    pf = db.execute(text("SELECT * FROM payment_files WHERE id=:id"), {"id": file_id}).fetchone()
    if not pf: raise HTTPException(404)
    pf = dict(pf._mapping)
    db.execute(text("DELETE FROM payment_transactions WHERE file_name=:fn AND portal=:p"), {"fn": pf["file_name"], "p": pf["portal"]})
    db.execute(text("DELETE FROM payment_files WHERE id=:id"), {"id": file_id})
    db.commit()
    return {"success": True}

@app.get("/api/portal-payment/summary")
def get_payment_summary(portal: str = None, month: str = None, year: int = None,
                        db: Session = Depends(get_db), current_user=Depends(auth.get_current_user)):
    params = {}
    w = "WHERE 1=1"
    if portal: w += " AND portal=:portal"; params["portal"] = portal
    if month:  w += " AND month=:month";  params["month"]  = month
    if year:   w += " AND year=:year";    params["year"]   = year
    row = db.execute(text(f"""SELECT SUM(total),SUM(product_sales),SUM(selling_fees),SUM(fba_fees),
        SUM(other_transaction_fees+other) FROM payment_transactions {w}"""), params).fetchone()
    breakdown = db.execute(text(f"""SELECT type,COUNT(*) as count,SUM(product_sales),SUM(selling_fees),
        SUM(fba_fees),SUM(other_transaction_fees+other),SUM(total)
        FROM payment_transactions {w} GROUP BY type ORDER BY SUM(total) DESC"""), params).fetchall()
    months = db.execute(text("SELECT DISTINCT month FROM payment_transactions WHERE month IS NOT NULL ORDER BY month")).fetchall()
    years  = db.execute(text("SELECT DISTINCT year  FROM payment_transactions WHERE year  IS NOT NULL ORDER BY year DESC")).fetchall()
    return {
        "total":         row[0] or 0,
        "product_sales": row[1] or 0,
        "selling_fees":  row[2] or 0,
        "fba_fees":      row[3] or 0,
        "other_fees":    row[4] or 0,
        "breakdown": [{"type":b[0],"count":b[1],"product_sales":b[2]or 0,"selling_fees":b[3]or 0,
                       "fba_fees":b[4]or 0,"other_fees":b[5]or 0,"total":b[6]or 0} for b in breakdown],
        "months": [r[0] for r in months],
        "years":  [r[0] for r in years],
    }

@app.get("/api/portal-payment/data")
def get_payment_data(portal: str = None, type: str = None, month: str = None, year: int = None,
                     search: str = None, page: int = 1, per_page: int = 50,
                     db: Session = Depends(get_db), current_user=Depends(auth.get_current_user)):
    q = "SELECT * FROM payment_transactions WHERE 1=1"
    params = {}
    if portal: q += " AND portal=:portal"; params["portal"] = portal
    if type:   q += " AND type=:type";     params["type"]   = type
    if month:  q += " AND month=:month";   params["month"]  = month
    if year:   q += " AND year=:year";     params["year"]   = year
    if search:
        q += " AND (sku LIKE :s OR order_id LIKE :s OR description LIKE :s)"; params["s"] = f"%{search}%"
    total = db.execute(text(q.replace("SELECT *","SELECT COUNT(*)")), params).scalar()
    q += f" ORDER BY date_time DESC LIMIT {per_page} OFFSET {(page-1)*per_page}"
    rows = db.execute(text(q), params).fetchall()
    return {"total": total, "page": page, "per_page": per_page, "items": [dict(r._mapping) for r in rows]}

@app.get("/api/portal-payment/export")
def export_payment_data(portal: str = None, type: str = None, month: str = None, year: int = None,
                        token: str = None, db: Session = Depends(get_db)):
    import csv as csvmod, io
    from fastapi.responses import StreamingResponse
    from jose import jwt
    SECRET_KEY = "rajnandini-fashion-secret-key-2024"
    try:
        payload = jwt.decode(token, SECRET_KEY, algorithms=["HS256"])
        if not payload.get("sub"): raise HTTPException(401)
    except: raise HTTPException(401)
    q = "SELECT * FROM payment_transactions WHERE 1=1"
    params = {}
    if portal: q += " AND portal=:portal"; params["portal"] = portal
    if type:   q += " AND type=:type";     params["type"]   = type
    if month:  q += " AND month=:month";   params["month"]  = month
    if year:   q += " AND year=:year";     params["year"]   = year
    q += " ORDER BY date_time DESC"
    rows = db.execute(text(q), params).fetchall()
    output = io.StringIO()
    w = csvmod.writer(output)
    w.writerow(["Date","Portal","Type","Order ID","SKU","Description","Product Sales","Selling Fees","FBA Fees","Other Fees","Other","Total","Fulfillment","Month","Year"])
    for r in rows:
        d = dict(r._mapping)
        w.writerow([d.get("date_time",""),d.get("portal",""),d.get("type",""),d.get("order_id",""),
                    d.get("sku",""),d.get("description",""),d.get("product_sales",0),d.get("selling_fees",0),
                    d.get("fba_fees",0),d.get("other_transaction_fees",0),d.get("other",0),d.get("total",0),
                    d.get("fulfillment",""),d.get("month",""),d.get("year","")])
    output.seek(0)
    from datetime import datetime
    fname = f"PortalPayment_{datetime.now().strftime('%Y%m%d_%H%M')}.csv"
    return StreamingResponse(iter([output.getvalue()]), media_type="text/csv",
        headers={"Content-Disposition": f"attachment; filename={fname}"})

@app.get("/api/items/export")
def export_items(search: str = None, sort: str = "sku", dir: str = "asc",
                 token: str = None,
                 db: Session = Depends(get_db)):
    import csv, io
    from fastapi.responses import StreamingResponse
    from sqlalchemy import text
    from jose import jwt, JWTError
    SECRET_KEY = "rajnandini-fashion-secret-key-2024"
    try:
        payload = jwt.decode(token, SECRET_KEY, algorithms=["HS256"])
        if not payload.get("sub"): raise HTTPException(401, "Invalid token")
    except: raise HTTPException(401, "Invalid token")
    allowed = {"sku","category","style_id","uniware_stock","fba_stock","sjit_stock","fbf_stock","location","mrp","cost_price"}
    sort_col = sort if sort in allowed else "sku"
    sort_dir = "DESC" if dir == "desc" else "ASC"
    q = "SELECT * FROM items WHERE 1=1"
    params = {}
    if search:
        q += " AND (sku LIKE :s OR category LIKE :s OR catalog_name LIKE :s OR style_id LIKE :s OR location LIKE :s)"
        params["s"] = f"%{search}%"
    q += f" ORDER BY {sort_col} {sort_dir}"
    rows = db.execute(text(q), params).fetchall()
    output = io.StringIO()
    w = csv.writer(output)
    w.writerow(["Uniware SKU","Category","Style ID","Catalog Name","Availability",
                "Uniware","FBA","SJIT","FBF","Location","Cost","Remark","Last Synced"])
    for r in rows:
        d = dict(r._mapping)
        w.writerow([d.get("sku",""), d.get("category",""), d.get("style_id",""),
                    d.get("catalog_name",""), d.get("availability",""),
                    d.get("uniware_stock",0), d.get("fba_stock",0),
                    d.get("sjit_stock",0), d.get("fbf_stock",0),
                    d.get("location",""), d.get("cost_price",0),
                    d.get("remark",""), d.get("synced_at","")])
    output.seek(0)
    from datetime import datetime
    fname = f"ItemMaster_{datetime.now().strftime('%Y%m%d_%H%M')}.csv"
    return StreamingResponse(iter([output.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": f"attachment; filename={fname}"})


# ─── PORTAL SALES ────────────────────────────────────────────────────────────

import os, re, glob
from pathlib import Path

def get_portal_sales_folder():
    username = os.environ.get("USERNAME") or os.environ.get("USER") or ""
    # Point directly to Amazon subfolder; scan month/year subfolders inside it
    candidates = [
        Path(f"C:/Users/{username}/OneDrive/Rajnandini/01 IMP_Do Not Edit  Delete/06 Portal Sales Files/Amazon"),
        Path(f"C:/Users/{username}/Rajnandini/01 IMP_Do Not Edit  Delete/06 Portal Sales Files/Amazon"),
    ]
    for p in candidates:
        if p.exists():
            return p
    # Return first candidate even if not exists (will show error in UI)
    return candidates[0]

def detect_amazon_meta(filename):
    """Detect month, year, report_type from Amazon MTR filename"""
    fn = filename.upper()
    report_type = "B2B" if "B2B" in fn else "B2C"
    months = {"JANUARY":1,"FEBRUARY":2,"MARCH":3,"APRIL":4,"MAY":5,"JUNE":6,
              "JULY":7,"AUGUST":8,"SEPTEMBER":9,"OCTOBER":10,"NOVEMBER":11,"DECEMBER":12}
    month_name = None
    for m in months:
        if m in fn:
            month_name = m  # store as UPPERCASE
            break
    # detect year
    year_match = re.search(r'20(\d{2})', filename)
    year = int("20" + year_match.group(1)) if year_match else None
    # custom report — extract from date range
    if month_name is None:
        date_match = re.search(r'(\d{4})_(\d{2})_(\d{2})', filename)
        if date_match:
            year = int(date_match.group(1))
            m_num = int(date_match.group(2))
            month_name = list(months.keys())[m_num-1]  # store as UPPERCASE
    return report_type, month_name, year

def parse_amazon_row(row, headers, report_type, portal, file_name, month, year):
    def g(col):
        col_lower = col.lower()
        for i, h in enumerate(headers):
            if h.lower() == col_lower:
                v = row[i] if i < len(row) else ""
                try: return v.strip()
                except: return ""
        return ""
    try: qty = int(float(g("Quantity") or 0))
    except: qty = 0
    try: inv_amt = float(g("Invoice Amount") or 0)
    except: inv_amt = 0
    try: teg = float(g("Tax Exclusive Gross") or 0)
    except: teg = 0
    # Normalize invoice date to DD-MM-YYYY
    raw_date = g("Invoice Date")
    try:
        from datetime import datetime as dt
        for fmt in ("%Y-%m-%d %H:%M:%S", "%d/%m/%y %H:%M", "%d/%m/%Y", "%Y-%m-%d"):
            try:
                d = dt.strptime(raw_date.strip(), fmt)
                raw_date = d.strftime("%d-%m-%Y")
                break
            except: pass
    except: pass
    return dict(
        portal=portal, report_type=report_type, file_name=file_name,
        seller_gstin=g("Seller Gstin"), invoice_number=g("Invoice Number"),
        invoice_date=raw_date, transaction_type=g("Transaction Type"),
        order_id=g("Order Id"), shipment_id=g("Shipment Id"),
        quantity=qty, asin=g("Asin"), sku=g("Sku"),
        item_description=g("Item Description"),
        fulfillment_channel=g("Fulfillment Channel"),
        ship_from_city=g("Ship From City"), ship_from_state=g("Ship From State"),
        ship_from_postal=g("Ship From Postal Code"),
        ship_to_city=g("Ship To City"), ship_to_state=g("Ship To State"),
        ship_to_postal=g("Ship To Postal Code"),
        invoice_amount=inv_amt, tax_exclusive_gross=teg,
        month=(month or "").upper().strip() or None, year=year, mapped_type=None, uniware_sku=None
    )

@app.get("/api/portal-sales/scan")
def scan_portal_folder(db: Session = Depends(get_db), current_user=Depends(auth.get_current_user)):
    try:
        from models import PortalFile
        from datetime import datetime
        ensure_portal_sale_tables(db)
        # Ensure portal_files table exists (ORM-based)
        try:
            db.execute(text("""CREATE TABLE IF NOT EXISTS portal_files (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                portal TEXT, file_name TEXT, file_path TEXT UNIQUE,
                report_type TEXT, month TEXT, year INTEGER,
                status TEXT DEFAULT 'pending', row_count INTEGER DEFAULT 0,
                detected_at DATETIME, loaded_at DATETIME
            )"""))
            db.commit()
        except: pass
        # Amazon folder — files live inside month/year subfolders
        amazon_folder = get_portal_sales_folder()
        portal = "Amazon"
        if not amazon_folder.exists():
            return {"error": f"Folder not found: {amazon_folder}", "files": [], "folder_path": str(amazon_folder)}
        new_files = []
        # Collect all files from subfolders (month/year) and root level
        all_found = list(amazon_folder.rglob("*.csv")) + list(amazon_folder.rglob("*.xlsx")) + list(amazon_folder.rglob("*.xls"))
        for f in all_found:
            try:
                exists = db.execute(
                    text("SELECT id FROM portal_files WHERE file_path=:p"),
                    {"p": str(f)}
                ).fetchone()
            except: exists = None
            if not exists:
                report_type, month, year = detect_amazon_meta(f.name)
                try:
                    pf = PortalFile(
                        portal=portal, file_name=f.name, file_path=str(f),
                        report_type=report_type, month=month, year=year,
                        status="pending", row_count=0,
                        detected_at=datetime.now()
                    )
                    db.add(pf)
                    db.flush()
                except Exception:
                    db.rollback()
                    try:
                        db.execute(text("""INSERT OR IGNORE INTO portal_files
                            (portal,file_name,file_path,report_type,month,year,status,row_count,detected_at)
                            VALUES (:po,:fn,:fp,:rt,:mo,:yr,'pending',0,:dt)"""),
                            {"po":portal,"fn":f.name,"fp":str(f),"rt":report_type,
                             "mo":month,"yr":year,"dt":datetime.now()})
                    except: pass
                new_files.append(f.name)
        db.commit()
        all_files = db.execute(text("""
            SELECT id, portal, file_name, report_type, month, year, status, row_count, detected_at, loaded_at
            FROM portal_files ORDER BY detected_at DESC
        """)).fetchall()
        return {"new": len(new_files), "files": [dict(r._mapping) for r in all_files], "folder_path": str(amazon_folder)}
    except Exception as e:
        import traceback
        return {"error": str(e), "detail": traceback.format_exc(), "files": [], "folder_path": ""}

@app.get("/api/portal-sales/files")
def get_portal_files(db: Session = Depends(get_db), current_user=Depends(auth.get_current_user)):
    rows = db.execute(text("""
        SELECT id, portal, file_name, report_type, month, year, status, row_count, detected_at, loaded_at
        FROM portal_files ORDER BY detected_at DESC
    """)).fetchall()
    return [dict(r._mapping) for r in rows]

@app.post("/api/portal-sales/load/{file_id}")
def load_portal_file(file_id: int, db: Session = Depends(get_db), current_user=Depends(auth.get_current_user)):
    import csv as csvmod
    from models import PortalFile, PortalSale
    from datetime import datetime
    pf = db.execute(text("SELECT * FROM portal_files WHERE id=:id"), {"id": file_id}).fetchone()
    if not pf: raise HTTPException(404, "File not found")
    pf = dict(pf._mapping)
    # delete existing rows for this file
    db.execute(text("DELETE FROM portal_sales WHERE file_name=:fn AND portal=:p"),
               {"fn": pf["file_name"], "p": pf["portal"]})
    count = 0
    try:
        fname = pf["file_path"]
        if fname.endswith(".xlsx") or fname.endswith(".xls"):
            import openpyxl
            wb = openpyxl.load_workbook(fname, read_only=True, data_only=True)
            ws = wb.active
            rows_iter = ws.iter_rows(values_only=True)
            headers = [str(h).strip() if h else "" for h in next(rows_iter)]
            for row in rows_iter:
                row = [str(c).strip() if c is not None else "" for c in row]
                data = parse_amazon_row(row, headers, pf["report_type"], pf["portal"],
                                        pf["file_name"], pf["month"], pf["year"])
                db.add(PortalSale(**data))
                count += 1
        else:
            with open(fname, encoding="utf-8-sig") as f:
                reader = csvmod.reader(f)
                headers = next(reader)
                for row in reader:
                    data = parse_amazon_row(row, headers, pf["report_type"], pf["portal"],
                                            pf["file_name"], pf["month"], pf["year"])
                    db.add(PortalSale(**data))
                    count += 1
        db.execute(text("UPDATE portal_files SET status='loaded', row_count=:c, loaded_at=:t WHERE id=:id"),
                   {"c": count, "t": datetime.now(), "id": file_id})
        db.commit()

        # ── Auto-map 1: Apply existing transaction type mappings ──
        mappings = db.execute(text("SELECT portal, transaction_type, mapped_to FROM transaction_type_mappings")).fetchall()
        for m in mappings:
            db.execute(text("UPDATE portal_sales SET mapped_type=:mt WHERE portal=:p AND transaction_type=:t AND file_name=:fn"),
                       {"mt": m.mapped_to, "p": m.portal, "t": m.transaction_type, "fn": pf["file_name"]})

        # ── Auto-map 2: Match portal SKU with Item Master SKU ──
        items = db.execute(text("SELECT sku FROM items")).fetchall()
        item_skus = {r.sku for r in items}
        # Also check existing sku_mappings
        sku_map = db.execute(text("SELECT portal_sku, item_master_sku FROM sku_mappings WHERE portal=:p"),
                             {"p": pf["portal"]}).fetchall()
        sku_dict = {r.portal_sku: r.item_master_sku for r in sku_map}

        portal_skus = db.execute(text("SELECT DISTINCT id, sku FROM portal_sales WHERE file_name=:fn AND portal=:p"),
                                 {"fn": pf["file_name"], "p": pf["portal"]}).fetchall()

        new_mappings = 0
        for row in portal_skus:
            psku = row.sku
            if not psku: continue
            mapped = None
            # Direct match in item master
            if psku in item_skus:
                mapped = psku
            # Existing sku_mapping
            elif psku in sku_dict:
                mapped = sku_dict[psku]
            if mapped:
                db.execute(text("UPDATE portal_sales SET uniware_sku=:u WHERE sku=:s AND portal=:p AND file_name=:fn"),
                           {"u": mapped, "s": psku, "p": pf["portal"], "fn": pf["file_name"]})
            else:
                # Add to sku_mappings as unmapped if not already there
                exists = db.execute(text("SELECT id FROM sku_mappings WHERE portal_sku=:s AND portal=:p"),
                                    {"s": psku, "p": pf["portal"]}).fetchone()
                if not exists:
                    db.execute(text("INSERT INTO sku_mappings (portal_sku, portal, item_master_sku) VALUES (:s,:p,NULL)"),
                               {"s": psku, "p": pf["portal"]})
                    new_mappings += 1

        db.commit()
        return {"success": True, "rows": count, "new_unmapped_skus": new_mappings}
    except Exception as e:
        db.rollback()
        raise HTTPException(500, str(e))

@app.get("/api/portal-sales/load-stream/{file_id}")
async def load_portal_file_stream(file_id: int, token: str, db: Session = Depends(get_db)):
    import csv as csvmod, json, asyncio
    from fastapi.responses import StreamingResponse
    from jose import jwt, JWTError
    from models import PortalFile, PortalSale
    from datetime import datetime
    SECRET_KEY = "rajnandini-fashion-secret-key-2024"
    try:
        payload = jwt.decode(token, SECRET_KEY, algorithms=["HS256"])
        username = payload.get("sub")
        if not username: raise HTTPException(401)
    except: raise HTTPException(401)

    async def stream():
        try:
            pf = db.execute(text("SELECT * FROM portal_files WHERE id=:id"), {"id": file_id}).fetchone()
            if not pf:
                yield f"data: {json.dumps({'stage':'error','msg':'File not found'})}\n\n"; return
            pf = dict(pf._mapping)
            fname = pf["file_path"]
            yield f"data: {json.dumps({'stage':'reading','pct':5,'msg':'Reading file...'})}\n\n"
            await asyncio.sleep(0)

            # Read file
            import concurrent.futures, openpyxl as oxl
            loop = asyncio.get_event_loop()
            def read_file():
                if fname.lower().endswith(('.xlsx','.xls')):
                    wb = oxl.load_workbook(fname, read_only=True, data_only=True)
                    ws = wb.active
                    rows = list(ws.iter_rows(values_only=True))
                    wb.close()
                    hdrs = [str(h).strip() if h else "" for h in rows[0]]
                    data = [[str(c).strip() if c is not None else "" for c in r] for r in rows[1:]]
                    return hdrs, data
                else:
                    with open(fname, encoding='utf-8-sig') as f:
                        reader = csvmod.reader(f)
                        hdrs = next(reader)
                        data = list(reader)
                    return hdrs, data
            with concurrent.futures.ThreadPoolExecutor() as pool:
                headers, rows_data = await loop.run_in_executor(pool, read_file)

            total = len(rows_data)
            yield f"data: {json.dumps({'stage':'parsing','pct':15,'total':total,'msg':f'Parsing {total} rows...'})}\n\n"
            await asyncio.sleep(0)

            # Delete old rows
            db.execute(text("DELETE FROM portal_sales WHERE file_name=:fn AND portal=:p"),
                       {"fn": pf["file_name"], "p": pf["portal"]})
            db.commit()

            # Insert in chunks
            CHUNK = 300
            inserted = 0
            from models import PortalSale
            for i in range(0, total, CHUNK):
                chunk = rows_data[i:i+CHUNK]
                for row in chunk:
                    try:
                        data = parse_amazon_row(row, headers, pf["report_type"], pf["portal"],
                                                pf["file_name"], pf["month"], pf["year"])
                        db.add(PortalSale(**data))
                        inserted += 1
                    except: pass
                db.commit()
                pct = 15 + int(((i+len(chunk))/total)*70)
                yield f"data: {json.dumps({'stage':'inserting','pct':pct,'inserted':inserted,'msg':f'Imported {inserted}/{total} rows'})}\n\n"
                await asyncio.sleep(0)

            # Update file status
            db.execute(text("UPDATE portal_files SET status='loaded', row_count=:c, loaded_at=:t WHERE id=:id"),
                       {"c": inserted, "t": datetime.now(), "id": file_id})
            db.commit()

            # ── Fast mapping phase (bulk SQL — no per-row loops) ────────────
            yield f"data: {json.dumps({'stage':'mapping','pct':88,'msg':'Applying mappings...'})}\n\n"
            await asyncio.sleep(0)

            fname_p  = pf["file_name"]
            portal_p = pf["portal"]

            # Ensure indexes for fast UPDATE/SELECT on this file's rows
            for _idx in [
                "CREATE INDEX IF NOT EXISTS idx_ps_fn_portal ON portal_sales (file_name, portal)",
                "CREATE INDEX IF NOT EXISTS idx_ps_txn       ON portal_sales (portal, transaction_type)",
                "CREATE INDEX IF NOT EXISTS idx_ps_sku       ON portal_sales (portal, sku)",
            ]:
                try: db.execute(text(_idx))
                except: pass
            db.commit()

            # 1) Transaction-type mappings — one CASE UPDATE instead of N loops
            _tmaps = db.execute(text(
                "SELECT transaction_type, mapped_to FROM transaction_type_mappings WHERE portal=:p"
            ), {"p": portal_p}).fetchall()
            if _tmaps:
                _case = " ".join(
                    "WHEN transaction_type=:tt{i} THEN :mt{i}".replace("{i}", str(i))
                    for i, _ in enumerate(_tmaps)
                )
                _params = {"fn": fname_p, "p": portal_p}
                for i, m in enumerate(_tmaps):
                    _params[f"tt{i}"] = m[0]
                    _params[f"mt{i}"] = m[1]
                db.execute(text(
                    f"UPDATE portal_sales SET mapped_type = CASE {_case} ELSE mapped_type END "
                    f"WHERE file_name=:fn AND portal=:p"
                ), _params)
                db.commit()

            # 2) SKU mapping — bulk UPDATE via subquery (no per-row loop)
            # Direct match: sku exists in items table
            db.execute(text(
                "UPDATE portal_sales SET uniware_sku=sku "
                "WHERE file_name=:fn AND portal=:p "
                "AND (uniware_sku IS NULL OR uniware_sku='') "
                "AND sku IN (SELECT sku FROM items)"
            ), {"fn": fname_p, "p": portal_p})

            # Mapped match: sku has an entry in sku_mappings
            db.execute(text(
                "UPDATE portal_sales SET uniware_sku=("
                "  SELECT sm.item_master_sku FROM sku_mappings sm"
                "  WHERE sm.portal_sku=portal_sales.sku AND sm.portal=portal_sales.portal"
                "  AND sm.item_master_sku IS NOT NULL AND sm.item_master_sku!=''"
                "  LIMIT 1"
                ") WHERE file_name=:fn AND portal=:p "
                "AND (uniware_sku IS NULL OR uniware_sku='') "
                "AND sku IN (SELECT portal_sku FROM sku_mappings WHERE portal=:p "
                "  AND item_master_sku IS NOT NULL AND item_master_sku!='')"
            ), {"fn": fname_p, "p": portal_p})
            db.commit()

            # 3) Register still-unmapped SKUs — one INSERT OR IGNORE for all
            db.execute(text(
                "INSERT OR IGNORE INTO sku_mappings (portal_sku, portal, item_master_sku) "
                "SELECT DISTINCT ps.sku, ps.portal, NULL FROM portal_sales ps "
                "WHERE ps.file_name=:fn AND ps.portal=:p "
                "AND ps.sku IS NOT NULL AND ps.sku!='' "
                "AND (ps.uniware_sku IS NULL OR ps.uniware_sku='') "
                "AND NOT EXISTS ("
                "  SELECT 1 FROM sku_mappings sm WHERE sm.portal_sku=ps.sku AND sm.portal=ps.portal"
                ")"
            ), {"fn": fname_p, "p": portal_p})
            db.commit()

            new_mappings = db.execute(text(
                "SELECT COUNT(*) FROM sku_mappings WHERE portal=:p AND (item_master_sku IS NULL OR item_master_sku='')"
            ), {"p": portal_p}).scalar() or 0
            db.commit()
            yield f"data: {json.dumps({'stage':'done','pct':100,'inserted':inserted,'new_unmapped_skus':new_mappings,'msg':f'Done! {inserted} rows loaded'})}\n\n"
        except Exception as e:
            import traceback
            yield f"data: {json.dumps({'stage':'error','msg':str(e)})}\n\n"

    return StreamingResponse(stream(), media_type="text/event-stream",
        headers={"Cache-Control":"no-cache","X-Accel-Buffering":"no"})

# ── AMAZON OTHER REPORT ──────────────────────────────────────────────────────

from pathlib import Path as _AOPath

def get_ao_base():
    import os
    username = os.environ.get("USERNAME") or os.environ.get("USER") or ""
    for base in [
        _AOPath(f"C:/Users/{username}/OneDrive/Rajnandini/01 IMP_Do Not Edit  Delete/07 Portal Payment File/Amazon"),
        _AOPath(f"C:/Users/{username}/Rajnandini/01 IMP_Do Not Edit  Delete/07 Portal Payment File/Amazon"),
    ]:
        if base.exists(): return base
    return _AOPath(f"C:/Users/{username}/OneDrive/Rajnandini/01 IMP_Do Not Edit  Delete/07 Portal Payment File/Amazon")

def get_ao_subfolders():
    """Dynamically return all subfolders from the Amazon path as {key: folder_name}.
    Key = folder name lowercased with spaces/special chars replaced by hyphens.
    Numbers are KEPT so keys are always unique and stable.
    """
    base = get_ao_base()
    if not base.exists():
        return {}
    import re as _re2
    folders = {}
    for d in sorted(base.iterdir()):
        if d.is_dir():
            # key: lowercase, spaces→hyphens, collapse multiple hyphens, trim
            key = d.name.lower().strip()
            key = _re2.sub(r'[^a-z0-9]+', '-', key).strip('-')
            folders[key] = d.name
    return folders

# For backward-compat: static fallback if folder doesn't exist yet
AO_FOLDERS_FALLBACK = {
    "payment":       "01 Payment Reports",
    "fba-return":    "02 FBA Customer Return",
    "replacement":   "03 Replacement",
    "reimbursement": "04 Reimbursements",
    "order-reimb":   "05 Order Level Reimbursement",
}

def get_ao_folders():
    """Get current subfolders, fall back to static list if path not found."""
    dynamic = get_ao_subfolders()
    return dynamic if dynamic else AO_FOLDERS_FALLBACK

@app.get("/api/amazon-other/tabs")
def ao_tabs(current_user=Depends(auth.get_current_user)):
    """Return list of subfolders from the Amazon path for dynamic tab building."""
    base = get_ao_base()
    folders = get_ao_folders()
    tabs = []
    for key, name in folders.items():
        folder_path = base / name
        tabs.append({
            "key": key,
            "name": name,
            "path": str(folder_path),
            "exists": folder_path.exists(),
        })
    return {"base": str(base), "base_exists": base.exists(), "tabs": tabs}

def ensure_ao_tables(db):
    db.execute(text("""CREATE TABLE IF NOT EXISTS ao_files (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        folder_key TEXT, file_name TEXT, file_path TEXT UNIQUE,
        period TEXT, status TEXT DEFAULT 'pending',
        row_count INTEGER DEFAULT 0,
        detected_at DATETIME, loaded_at DATETIME
    )"""))
    db.execute(text("""CREATE TABLE IF NOT EXISTS ao_rows (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        file_id INTEGER, folder_key TEXT, period TEXT,
        col1 TEXT, col2 TEXT, col3 TEXT, col4 TEXT, col5 TEXT,
        col6 TEXT, col7 TEXT, col8 TEXT, col9 TEXT, col10 TEXT,
        col11 TEXT, col12 TEXT, col13 TEXT, col14 TEXT, col15 TEXT, col16 TEXT,
        raw_json TEXT,
        -- pre-parsed numeric columns for fast SQL SUM (Amazon payment CSV)
        txn_type TEXT,
        product_sales REAL DEFAULT 0,
        shipping_credits REAL DEFAULT 0,
        promo_rebates REAL DEFAULT 0,
        tcs_cgst REAL DEFAULT 0,
        tcs_sgst REAL DEFAULT 0,
        tcs_igst REAL DEFAULT 0,
        tds REAL DEFAULT 0,
        selling_fees REAL DEFAULT 0,
        fba_fees REAL DEFAULT 0,
        other_txn_fees REAL DEFAULT 0,
        other_amt REAL DEFAULT 0,
        total_amt REAL DEFAULT 0
    )"""))
    db.commit()
    # Migrations: add columns that may be missing if table existed before this version
    for tbl, cols in [
        ("ao_files", [("folder_key","TEXT"),("period","TEXT"),("loaded_at","DATETIME")]),
        ("ao_rows",  [("folder_key","TEXT"),("period","TEXT"),("col13","TEXT"),
                      ("col14","TEXT"),("col15","TEXT"),("col16","TEXT"),("raw_json","TEXT"),
                      ("txn_type","TEXT"),
                      ("product_sales","REAL DEFAULT 0"),("shipping_credits","REAL DEFAULT 0"),
                      ("promo_rebates","REAL DEFAULT 0"),("tcs_cgst","REAL DEFAULT 0"),
                      ("tcs_sgst","REAL DEFAULT 0"),("tcs_igst","REAL DEFAULT 0"),
                      ("tds","REAL DEFAULT 0"),("selling_fees","REAL DEFAULT 0"),
                      ("fba_fees","REAL DEFAULT 0"),("other_txn_fees","REAL DEFAULT 0"),
                      ("other_amt","REAL DEFAULT 0"),("total_amt","REAL DEFAULT 0")]),
    ]:
        for col, cdef in cols:
            try:
                db.execute(text(f"ALTER TABLE {tbl} ADD COLUMN {col} {cdef}"))
                db.commit()
            except: pass
    # Backfill runs separately via /api/amazon-other/backfill endpoint

def _ao_backfill_numeric(db):
    """Parse raw_json and populate numeric columns for rows not yet backfilled."""
    import json as _j
    # Only process rows where total_amt is NULL or 0 AND raw_json has content
    rows = db.execute(text("""
        SELECT id, raw_json FROM ao_rows
        WHERE (total_amt IS NULL OR total_amt=0)
        AND raw_json IS NOT NULL AND raw_json!='{}'
        AND raw_json LIKE '%product sales%'
        LIMIT 50000
    """)).fetchall()
    if not rows:
        return
    def _fv(d, *keys):
        for k in keys:
            for dk, dv in d.items():
                if dk.strip().lower() == k.lower():
                    try: return float(str(dv).replace(',','').strip() or 0)
                    except: return 0.0
        return 0.0
    CHUNK = 500
    for i in range(0, len(rows), CHUNK):
        batch = rows[i:i+CHUNK]
        for row in batch:
            try:
                d = _j.loads(row[1])
                db.execute(text("""UPDATE ao_rows SET
                    txn_type=:tt,
                    product_sales=:ps, shipping_credits=:sc, promo_rebates=:pr,
                    tcs_cgst=:tc1, tcs_sgst=:tc2, tcs_igst=:tc3, tds=:td,
                    selling_fees=:sf, fba_fees=:fba, other_txn_fees=:otf,
                    other_amt=:oth, total_amt=:tot
                    WHERE id=:id"""),
                    {"id": row[0],
                     "tt":  d.get("type","").strip(),
                     "ps":  _fv(d,"product sales"),
                     "sc":  _fv(d,"shipping credits"),
                     "pr":  _fv(d,"promotional rebates"),
                     "tc1": _fv(d,"TCS-CGST"),
                     "tc2": _fv(d,"TCS-SGST"),
                     "tc3": _fv(d,"TCS-IGST"),
                     "td":  _fv(d,"TDS (Section 194-O)"),
                     "sf":  _fv(d,"selling fees"),
                     "fba": _fv(d,"fba fees"),
                     "otf": _fv(d,"other transaction fees"),
                     "oth": _fv(d,"other"),
                     "tot": _fv(d,"total")})
            except: pass
        db.commit()

@app.get("/api/amazon-other/scan-all")
def ao_scan_all(db: Session = Depends(get_db), current_user=Depends(auth.get_current_user)):
    """Scan ALL subfolders of the Amazon path in one call. Returns all files."""
    import re as _re
    from datetime import datetime

    def extract_period(filename):
        stem = _re.sub(r'\.(csv|xlsx|xls)$', '', filename, flags=_re.IGNORECASE).strip()
        # DD-MM-YYYY to DD-MM-YYYY
        m = _re.search(r'(\d{1,2}[-/]\d{1,2}[-/]\d{4})\s+to\s+(\d{1,2}[-/]\d{1,2}[-/]\d{4})', stem, _re.IGNORECASE)
        if m: return m.group(1) + ' to ' + m.group(2)
        # YYYY-MM-DD to YYYY-MM-DD
        m = _re.search(r'(\d{4}[-/]\d{1,2}[-/]\d{1,2})\s+to\s+(\d{4}[-/]\d{1,2}[-/]\d{1,2})', stem, _re.IGNORECASE)
        if m: return m.group(1) + ' to ' + m.group(2)
        # Month YYYY
        MONTHS = r'(jan(?:uary)?|feb(?:ruary)?|mar(?:ch)?|apr(?:il)?|may|jun(?:e)?|jul(?:y)?|aug(?:ust)?|sep(?:tember)?|oct(?:ober)?|nov(?:ember)?|dec(?:ember)?)'
        m = _re.search(MONTHS + r'[-\s]+(\d{4})', stem, _re.IGNORECASE)
        if m: return m.group(1).capitalize() + ' ' + m.group(2)
        # YYYY-MM
        m = _re.search(r'(20\d{2})[-_](\d{2})\b', stem)
        if m:
            try:
                import calendar
                return calendar.month_name[int(m.group(2))] + ' ' + m.group(1)
            except: return m.group(1) + '-' + m.group(2)
        # YYYY only
        m = _re.search(r'\b(20\d{2})\b', stem)
        if m: return m.group(1)
        return stem

    try:
        ensure_ao_tables(db)
        base = get_ao_base()
        if not base.exists():
            return {"error": f"Base folder not found: {base}", "new": 0, "files": [], "folder_path": str(base)}
        folders = get_ao_folders()
        new_files = []
        errors = []
        for folder_key, folder_name in folders.items():
            folder = base / folder_name
            if not folder.exists():
                errors.append(f"{folder_name}: not found")
                continue
            all_found = (list(folder.rglob("*.csv")) + list(folder.rglob("*.CSV")) +
                         list(folder.rglob("*.xlsx")) + list(folder.rglob("*.xls")))
            for f in all_found:
                try:
                    period = extract_period(f.name)
                    exists = db.execute(text("SELECT id FROM ao_files WHERE file_path=:p"), {"p": str(f)}).fetchone()
                    if not exists:
                        db.execute(text("""INSERT OR IGNORE INTO ao_files
                            (folder_key,file_name,file_path,period,status,row_count,detected_at)
                            VALUES (:fk,:fn,:fp,:pe,'pending',0,:dt)"""),
                            {"fk": folder_key, "fn": f.name, "fp": str(f), "pe": period, "dt": datetime.now()})
                        new_files.append(f"{folder_name}/{f.name}")
                    else:
                        # Fix period for existing entries that used filename as period
                        db.execute(text("UPDATE ao_files SET period=:pe WHERE file_path=:fp AND (period IS NULL OR period='' OR period=file_name)"),
                                   {"pe": period, "fp": str(f)})
                except Exception as fe:
                    errors.append(str(fe))
        # Remove stale DB entries
        all_db = db.execute(text("SELECT id, file_path FROM ao_files")).fetchall()
        from pathlib import Path as _P
        for row in all_db:
            if not _P(row[1]).exists():
                db.execute(text("DELETE FROM ao_rows WHERE file_id=:id"), {"id": row[0]})
                db.execute(text("DELETE FROM ao_files WHERE id=:id"), {"id": row[0]})
        db.commit()
        rows = db.execute(text("SELECT * FROM ao_files ORDER BY folder_key, period DESC, file_name")).fetchall()
        return {"new": len(new_files), "files": [dict(r._mapping) for r in rows],
                "folder_path": str(base), "errors": errors, "scanned_folders": list(folders.keys())}
    except Exception as e:
        import traceback
        return {"error": str(e), "detail": traceback.format_exc(), "new": 0, "files": [], "folder_path": ""}

@app.get("/api/amazon-other/scan/{folder_key}")
def ao_scan(folder_key: str, db: Session = Depends(get_db), current_user=Depends(auth.get_current_user)):
    try:
        ensure_ao_tables(db)
        folders = get_ao_folders()
        if folder_key not in folders:
            return {"error": f"Unknown folder key: {folder_key}", "files": [], "folder_path": ""}
        base = get_ao_base()
        folder = base / folders[folder_key]
        if not folder.exists():
            return {"error": f"Folder not found: {folder}", "files": [], "folder_path": str(folder)}
        from datetime import datetime
        new_files = []
        all_found = list(folder.rglob("*.csv")) + list(folder.rglob("*.xlsx")) + list(folder.rglob("*.xls"))
        for f in all_found:
            exists = db.execute(text("SELECT id FROM ao_files WHERE file_path=:p"), {"p": str(f)}).fetchone()
            if not exists:
                period = f.parent.name if f.parent != folder else ""
                db.execute(text("""INSERT OR IGNORE INTO ao_files
                    (folder_key,file_name,file_path,period,status,row_count,detected_at)
                    VALUES (:fk,:fn,:fp,:pe,'pending',0,:dt)"""),
                    {"fk": folder_key, "fn": f.name, "fp": str(f), "pe": period, "dt": datetime.now()})
                new_files.append(f.name)
        db.commit()
        rows = db.execute(text("SELECT * FROM ao_files WHERE folder_key=:fk ORDER BY period DESC, file_name"), {"fk": folder_key}).fetchall()
        return {"new": len(new_files), "files": [dict(r._mapping) for r in rows], "folder_path": str(folder)}
    except Exception as e:
        import traceback
        return {"error": str(e), "files": [], "folder_path": ""}

@app.get("/api/amazon-other/backfill")
def ao_backfill(db: Session = Depends(get_db), current_user=Depends(auth.get_current_user)):
    """Backfill numeric cols from raw_json for rows loaded before the schema upgrade."""
    try:
        ensure_ao_tables(db)
        # Count rows needing backfill
        needs = db.execute(text("""
            SELECT COUNT(*) FROM ao_rows
            WHERE (total_amt IS NULL OR total_amt=0)
            AND raw_json IS NOT NULL AND raw_json NOT IN ('{}','')
            AND raw_json LIKE '%product sales%'
        """)).scalar() or 0
        if needs == 0:
            return {"backfilled": 0, "msg": "Nothing to backfill"}
        _ao_backfill_numeric(db)
        return {"backfilled": needs, "msg": f"Backfilled {needs} rows"}
    except Exception as e:
        return {"backfilled": 0, "error": str(e)}

@app.get("/api/amazon-other/files-all")
def ao_files_all(db: Session = Depends(get_db), current_user=Depends(auth.get_current_user)):
    """Return all files from all folders in one query."""
    try:
        ensure_ao_tables(db)
        rows = db.execute(text(
            "SELECT * FROM ao_files ORDER BY folder_key, period DESC, file_name"
        )).fetchall()
        return [dict(r._mapping) for r in rows]
    except: return []

@app.get("/api/amazon-other/files/{folder_key}")
def ao_files(folder_key: str, db: Session = Depends(get_db), current_user=Depends(auth.get_current_user)):
    try:
        ensure_ao_tables(db)
        rows = db.execute(text("SELECT * FROM ao_files WHERE folder_key=:fk ORDER BY period DESC, file_name"), {"fk": folder_key}).fetchall()
        return [dict(r._mapping) for r in rows]
    except: return []

@app.get("/api/amazon-other/load-stream/{file_id}")
async def ao_load_stream(file_id: int, token: str, folder_key: str = "", db: Session = Depends(get_db)):
    import csv as csvmod, json, asyncio, concurrent.futures
    from fastapi.responses import StreamingResponse
    from jose import jwt as _jwt
    from datetime import datetime
    SECRET_KEY = "rajnandini-fashion-secret-key-2024"
    try:
        payload = _jwt.decode(token, SECRET_KEY, algorithms=["HS256"])
        if not payload.get("sub"): raise HTTPException(401)
    except: raise HTTPException(401)

    async def stream():
        try:
            ensure_ao_tables(db)
            pf = db.execute(text("SELECT * FROM ao_files WHERE id=:id"), {"id": file_id}).fetchone()
            if not pf:
                yield f"data: {json.dumps({'stage':'error','msg':'File not found'})}\n\n"; return
            pf = dict(pf._mapping)
            fkey = pf.get("folder_key") or folder_key
            yield f"data: {json.dumps({'stage':'reading','pct':5,'msg':'Reading file...'})}\n\n"
            await asyncio.sleep(0)
            fname = pf["file_path"]

            def read_file():
                """Smart reader: skips Amazon metadata preamble rows, finds real header."""
                import re as _re
                # Known real header columns for Amazon payment CSV
                HEADER_SIGNALS = {'date/time', 'settlement id', 'type', 'order id'}

                def find_header_row(all_rows):
                    """Return (header_idx, rows) where header_idx is the real header."""
                    for i, row in enumerate(all_rows):
                        row_lower = {str(c).strip().lower() for c in row if c}
                        if len(HEADER_SIGNALS & row_lower) >= 3:
                            return i
                    return 0  # fallback to first row

                if fname.lower().endswith(('.xlsx','.xls')):
                    import openpyxl
                    wb = openpyxl.load_workbook(fname, read_only=True, data_only=True)
                    ws = wb.active
                    all_rows = list(ws.iter_rows(values_only=True))
                    wb.close()
                    if not all_rows: return [], []
                    hi = find_header_row(all_rows)
                    hdrs = [str(h).strip() if h else "" for h in all_rows[hi]]
                    data = [[str(c).strip() if c is not None else "" for c in r]
                            for r in all_rows[hi+1:] if any(c is not None for c in r)]
                    return hdrs, data
                else:
                    with open(fname, encoding='utf-8-sig') as f:
                        all_rows = list(csvmod.reader(f))
                    if not all_rows: return [], []
                    hi = find_header_row(all_rows)
                    hdrs = [h.strip() for h in all_rows[hi]]
                    data = [r for r in all_rows[hi+1:] if any(c.strip() for c in r)]
                    return hdrs, data

            loop = asyncio.get_event_loop()
            with concurrent.futures.ThreadPoolExecutor() as pool:
                headers, rows_data = await loop.run_in_executor(pool, read_file)

            total = len(rows_data)
            yield f"data: {json.dumps({'stage':'parsing','pct':15,'total':total,'msg':f'Parsing {total} rows...'})}\n\n"
            await asyncio.sleep(0)

            db.execute(text("DELETE FROM ao_rows WHERE file_id=:fid"), {"fid": file_id})
            db.commit()

            CHUNK = 300
            inserted = 0
            import json as _json

            for i in range(0, total, CHUNK):
                chunk = rows_data[i:i+CHUNK]
                for row in chunk:
                    padded = (row + [''] * 16)[:16]
                    row_dict = dict(zip(headers[:len(row)], row)) if headers else {}

                    # Extract numeric fields for fast SQL SUM (handles Amazon payment CSV)
                    def _fv(key):
                        for k, v in row_dict.items():
                            if k.strip().lower() == key.lower():
                                try: return float(str(v).replace(',','').strip() or 0)
                                except: return 0.0
                        return 0.0

                    db.execute(text("""INSERT INTO ao_rows
                        (file_id,folder_key,period,col1,col2,col3,col4,col5,col6,col7,col8,
                         col9,col10,col11,col12,col13,col14,col15,col16,raw_json,
                         txn_type,product_sales,shipping_credits,promo_rebates,
                         tcs_cgst,tcs_sgst,tcs_igst,tds,selling_fees,fba_fees,
                         other_txn_fees,other_amt,total_amt)
                        VALUES (:fid,:fk,:pe,:c1,:c2,:c3,:c4,:c5,:c6,:c7,:c8,
                                :c9,:c10,:c11,:c12,:c13,:c14,:c15,:c16,:rj,
                                :tt,:ps,:sc,:pr,:tc1,:tc2,:tc3,:td,:sf,:fba,:otf,:oth,:tot)"""),
                        {"fid":file_id,"fk":fkey,"pe":pf.get("period",""),
                         "c1":padded[0],"c2":padded[1],"c3":padded[2],"c4":padded[3],
                         "c5":padded[4],"c6":padded[5],"c7":padded[6],"c8":padded[7],
                         "c9":padded[8],"c10":padded[9],"c11":padded[10],"c12":padded[11],
                         "c13":padded[12],"c14":padded[13],"c15":padded[14],"c16":padded[15],
                         "rj":_json.dumps(row_dict),
                         "tt": row_dict.get("type","").strip(),
                         "ps":  _fv("product sales"),
                         "sc":  _fv("shipping credits"),
                         "pr":  _fv("promotional rebates"),
                         "tc1": _fv("TCS-CGST"),
                         "tc2": _fv("TCS-SGST"),
                         "tc3": _fv("TCS-IGST"),
                         "td":  _fv("TDS (Section 194-O)"),
                         "sf":  _fv("selling fees"),
                         "fba": _fv("fba fees"),
                         "otf": _fv("other transaction fees"),
                         "oth": _fv("other"),
                         "tot": _fv("total")})
                    inserted += 1
                db.commit()
                pct = 15 + int(((i+len(chunk))/total)*78)
                yield f"data: {json.dumps({'stage':'inserting','pct':pct,'inserted':inserted,'msg':f'Imported {inserted}/{total} rows'})}\n\n"
                await asyncio.sleep(0)

            db.execute(text("UPDATE ao_files SET status='loaded',row_count=:c,loaded_at=:t WHERE id=:id"),
                       {"c":inserted,"t":datetime.now(),"id":file_id})
            db.commit()
            yield f"data: {json.dumps({'stage':'done','pct':100,'inserted':inserted,'msg':f'Done! {inserted} rows loaded'})}\n\n"
        except Exception as e:
            import traceback
            yield f"data: {json.dumps({'stage':'error','msg':str(e)})}\n\n"

    return StreamingResponse(stream(), media_type="text/event-stream",
        headers={"Cache-Control":"no-cache","X-Accel-Buffering":"no"})

@app.delete("/api/amazon-other/delete/{file_id}")
def ao_delete(file_id: int, db: Session = Depends(get_db), current_user=Depends(auth.get_current_user)):
    ensure_ao_tables(db)
    db.execute(text("DELETE FROM ao_rows WHERE file_id=:fid"), {"fid": file_id})
    db.execute(text("DELETE FROM ao_files WHERE id=:id"), {"id": file_id})
    db.commit()
    return {"ok": True}

@app.get("/api/amazon-other/payment-summary/{folder_key}")
def ao_payment_summary(folder_key: str, period: str = None, txn_type: str = None,
                        date_from: str = None, date_to: str = None,
                        db: Session = Depends(get_db), current_user=Depends(auth.get_current_user)):
    """Fast KPI summary using pre-stored numeric columns — single SQL query."""
    try:
        ensure_ao_tables(db)
        conds = ["folder_key=:fk"]
        params = {"fk": folder_key}
        if period:   conds.append("period=:pe");      params["pe"] = period
        if txn_type: conds.append("txn_type=:tt");    params["tt"] = txn_type
        where = "WHERE " + " AND ".join(conds)

        # Single aggregation query — instant even on 200K rows
        row = db.execute(text(f"""
            SELECT
                COALESCE(SUM(product_sales),0)    AS product_sales,
                COALESCE(SUM(shipping_credits),0) AS shipping_credits,
                COALESCE(SUM(promo_rebates),0)    AS promo_rebates,
                COALESCE(SUM(tcs_cgst),0)+COALESCE(SUM(tcs_sgst),0)+COALESCE(SUM(tcs_igst),0) AS tcs,
                COALESCE(SUM(tds),0)              AS tds,
                COALESCE(SUM(selling_fees),0)     AS selling_fees,
                COALESCE(SUM(fba_fees),0)         AS fba_fees,
                COALESCE(SUM(other_txn_fees),0)   AS other_txn_fees,
                COALESCE(SUM(other_amt),0)        AS other_amt,
                COALESCE(SUM(total_amt),0)        AS total_amt,
                COUNT(*)                          AS total_rows,
                SUM(CASE WHEN txn_type='Order'  THEN 1 ELSE 0 END) AS order_count,
                SUM(CASE WHEN txn_type='Refund' THEN 1 ELSE 0 END) AS refund_count
            FROM ao_rows {where}
        """), params).fetchone()

        # Per-type breakdown — also fast SQL
        type_rows = db.execute(text(f"""
            SELECT txn_type,
                COUNT(*) as cnt,
                COALESCE(SUM(product_sales),0) as ps,
                COALESCE(SUM(selling_fees),0)  as sf,
                COALESCE(SUM(fba_fees),0)      as fba,
                COALESCE(SUM(total_amt),0)     as tot
            FROM ao_rows {where} AND txn_type IS NOT NULL AND txn_type!=''
            GROUP BY txn_type ORDER BY cnt DESC
        """), params).fetchall()

        def rnd(v): return round(float(v or 0), 2)
        r = dict(row._mapping) if row else {}
        summary = {
            "product_sales":   rnd(r.get("product_sales",0)),
            "shipping_credits":rnd(r.get("shipping_credits",0)),
            "promo_rebates":   rnd(r.get("promo_rebates",0)),
            "tcs":             rnd(r.get("tcs",0)),
            "tds":             rnd(r.get("tds",0)),
            "selling_fees":    rnd(r.get("selling_fees",0)),
            "fba_fees":        rnd(r.get("fba_fees",0)),
            "other_txn_fees":  rnd(r.get("other_txn_fees",0)),
            "other":           rnd(r.get("other_amt",0)),
            "total":           rnd(r.get("total_amt",0)),
            "rows":            int(r.get("total_rows",0)),
            "order_count":     int(r.get("order_count",0)),
            "refund_count":    int(r.get("refund_count",0)),
        }
        by_type = {}
        types = []
        for tr in type_rows:
            t = tr[0]
            types.append(t)
            by_type[t] = {"count":tr[1],"product_sales":rnd(tr[2]),"selling_fees":rnd(tr[3]),"fba_fees":rnd(tr[4]),"total":rnd(tr[5])}

        return {"summary": summary, "by_type": by_type, "types": types}
    except Exception as e:
        # Fallback: if numeric cols not yet populated (old data), return zeros quickly
        return {"summary": {"product_sales":0,"selling_fees":0,"fba_fees":0,"tcs":0,"tds":0,
                            "total":0,"rows":0,"order_count":0,"refund_count":0},
                "by_type": {}, "types": [], "error": str(e)}

@app.get("/api/amazon-other/filters/{folder_key}")
def ao_filters(folder_key: str, db: Session = Depends(get_db), current_user=Depends(auth.get_current_user)):
    """Return distinct years, months, periods for filter dropdowns."""
    try:
        ensure_ao_tables(db)
        periods = [r[0] for r in db.execute(text(
            "SELECT DISTINCT period FROM ao_rows WHERE folder_key=:fk AND period IS NOT NULL AND period!='' ORDER BY period DESC"),
            {"fk": folder_key}).fetchall()]
        # Try to extract year/month from period strings like "2026-01" or "January 2026"
        import re as _re
        years, months = set(), set()
        MN = ["January","February","March","April","May","June",
              "July","August","September","October","November","December"]
        for p in periods:
            m = _re.search(r'(20\d\d)', p)
            if m: years.add(m.group(1))
            for mn in MN:
                if mn.lower() in p.lower(): months.add(mn)
            # Handle YYYY-MM format
            m2 = _re.match(r'20\d\d-(\d\d)', p)
            if m2:
                try: months.add(MN[int(m2.group(1))-1])
                except: pass
        return {
            "periods": periods,
            "years": sorted(years, reverse=True),
            "months": [m for m in MN if m in months],
        }
    except Exception as e:
        return {"periods": [], "years": [], "months": [], "error": str(e)}

@app.get("/api/amazon-other/data/{folder_key}")
def ao_data(folder_key: str, page: int = 1, per_page: int = 10,
            search: str = None, period: str = None,
            year: str = None, month: str = None,
            date_from: str = None, date_to: str = None,
            txn_type: str = None,
            db: Session = Depends(get_db), current_user=Depends(auth.get_current_user)):
    try:
        import json as _json, re as _re
        ensure_ao_tables(db)
        conds = ["r.folder_key=:fk"]
        params = {"fk": folder_key}
        if period:
            conds.append("r.period=:pe"); params["pe"] = period
        if txn_type:
            # Filter by 'type' field stored in raw_json col3 (3rd column in Amazon payment CSV)
            conds.append("r.raw_json LIKE :tt")
            params["tt"] = f'%"type": "{txn_type}"%'
        if search:
            conds.append("(r.col1 LIKE :s OR r.col2 LIKE :s OR r.col3 LIKE :s OR r.col4 LIKE :s OR r.raw_json LIKE :s)")
            params["s"] = f"%{search}%"
        # year/month/date filters applied post-fetch on raw_json (flexible for any date column)
        where = "WHERE " + " AND ".join(conds)
        total_raw = db.execute(text(f"SELECT COUNT(*) FROM ao_rows r {where}"), params).scalar() or 0
        # Fetch all matching for client-side year/month/date filter (paginate after)
        if year or month or date_from or date_to:
            all_rows = db.execute(text(f"SELECT * FROM ao_rows r {where} ORDER BY r.id"), params).fetchall()
            items_all = []
            MN = ["january","february","march","april","may","june",
                  "july","august","september","october","november","december"]
            for r in all_rows:
                rd = dict(r._mapping)
                try: rd['data'] = _json.loads(rd.get('raw_json') or '{}')
                except: rd['data'] = {}
                # Find any date-like value in data
                date_str = ""
                for k, v in rd['data'].items():
                    if any(x in k.lower() for x in ['date','month','period']) and v:
                        date_str = str(v); break
                if not date_str: date_str = rd.get('period','')
                # Year filter
                if year and year not in date_str: continue
                # Month filter
                if month and month.lower() not in date_str.lower(): continue
                # Date range filter
                if date_from or date_to:
                    m = _re.search(r'(\d{4}-\d{2}-\d{2})', date_str)
                    if m:
                        ds = m.group(1)
                        if date_from and ds < date_from: continue
                        if date_to   and ds > date_to:   continue
                items_all.append(rd)
            total = len(items_all)
            offset = (page-1)*per_page
            items = items_all[offset:offset+per_page]
        else:
            total = total_raw
            offset = (page-1)*per_page
            rows = db.execute(text(f"SELECT * FROM ao_rows r {where} ORDER BY r.id LIMIT :lim OFFSET :off"),
                              {**params, "lim": per_page, "off": offset}).fetchall()
            items = []
            for r in rows:
                rd = dict(r._mapping)
                try: rd['data'] = _json.loads(rd.get('raw_json') or '{}')
                except: rd['data'] = {}
                items.append(rd)
        # Total amount if col has amount-like key
        total_amount = None
        try:
            total_amount = round(db.execute(text(f"SELECT SUM(CAST(REPLACE(REPLACE(col10,'₹',''),',','') AS REAL)) FROM ao_rows r {where}"), params).scalar() or 0, 2)
        except: pass
        return {"total": total, "items": items, "total_amount": total_amount}
    except Exception as e:
        return {"total": 0, "items": [], "error": str(e)}

@app.get("/api/amazon-other/export/{folder_key}")
def ao_export(folder_key: str, token: str, year: str = None, month: str = None,
              period: str = None, date_from: str = None, date_to: str = None,
              search: str = None, db: Session = Depends(get_db)):
    import csv as csvmod, io, json as _json
    from fastapi.responses import StreamingResponse
    from jose import jwt as _jwt
    SECRET_KEY = "rajnandini-fashion-secret-key-2024"
    try:
        payload = _jwt.decode(token, SECRET_KEY, algorithms=["HS256"])
        if not payload.get("sub"): raise HTTPException(401)
    except: raise HTTPException(401)
    d = ao_data(folder_key, page=1, per_page=999999, search=search,
                period=period, year=year, month=month,
                date_from=date_from, date_to=date_to, db=db, current_user=None)
    output = io.StringIO()
    w = csvmod.writer(output)
    if d["items"]:
        keys = list(d["items"][0].get('data',{}).keys()) or [f"col{i}" for i in range(1,17)]
        w.writerow(keys)
        for item in d["items"]:
            row_data = item.get('data',{})
            w.writerow([row_data.get(k,'') for k in keys])
    output.seek(0)
    from datetime import datetime
    return StreamingResponse(iter([output.getvalue()]), media_type="text/csv",
        headers={"Content-Disposition": f"attachment; filename=ao_{folder_key}_{datetime.now().strftime('%Y%m%d')}.csv"})

@app.get("/api/portal-sales/transaction-types")
def get_transaction_types(portal: str = None, db: Session = Depends(get_db), current_user=Depends(auth.get_current_user)):
    q = "SELECT DISTINCT portal, transaction_type FROM portal_sales WHERE 1=1"
    params = {}
    if portal:
        q += " AND portal=:p"; params["p"] = portal
    rows = db.execute(text(q), params).fetchall()
    types = [dict(r._mapping) for r in rows]
    # get existing mappings
    mappings = db.execute(text("SELECT portal, transaction_type, mapped_to FROM transaction_type_mappings")).fetchall()
    mapping_dict = {(m.portal, m.transaction_type): m.mapped_to for m in mappings}
    for t in types:
        t["mapped_to"] = mapping_dict.get((t["portal"], t["transaction_type"]))
    return types

@app.post("/api/portal-sales/transaction-types/map")
def map_transaction_type(data: dict, db: Session = Depends(get_db), current_user=Depends(auth.get_current_user)):
    from datetime import datetime
    portal = data["portal"]; tt = data["transaction_type"]; mapped = data["mapped_to"]
    exists = db.execute(text("SELECT id FROM transaction_type_mappings WHERE portal=:p AND transaction_type=:t"),
                        {"p": portal, "t": tt}).fetchone()
    if exists:
        db.execute(text("UPDATE transaction_type_mappings SET mapped_to=:m, updated_at=:u WHERE portal=:p AND transaction_type=:t"),
                   {"m": mapped, "u": datetime.now(), "p": portal, "t": tt})
    else:
        db.execute(text("INSERT INTO transaction_type_mappings (portal, transaction_type, mapped_to, updated_at) VALUES (:p,:t,:m,:u)"),
                   {"p": portal, "t": tt, "m": mapped, "u": datetime.now()})
    # update mapped_type in portal_sales
    db.execute(text("UPDATE portal_sales SET mapped_type=:m WHERE portal=:p AND transaction_type=:t"),
               {"m": mapped, "p": portal, "t": tt})
    db.commit()
    return {"success": True}

@app.get("/api/portal-sales/data")
def get_portal_sales_data(
    portal: str = None, month: str = None, year: str = None,
    mapped_type: str = None, search: str = None, unmapped_sku: int = 0,
    page: int = 1, per_page: int = 50,
    db: Session = Depends(get_db), current_user=Depends(auth.get_current_user)
):
    q = """SELECT ps.*, i.category
           FROM portal_sales ps
           LEFT JOIN items i ON ps.uniware_sku = i.sku
           WHERE 1=1"""
    params = {}
    if portal and portal not in ("", "ALL"):
        q += " AND ps.portal=:portal"; params["portal"] = portal
    if month and month not in ("", "ALL"):
        q += " AND ps.month=:month"; params["month"] = month.upper().strip()
    if year and str(year) not in ("", "ALL", "0"):
        q += " AND ps.year=:year"; params["year"] = int(year)
    if mapped_type and mapped_type not in ("", "ALL"):
        q += " AND ps.mapped_type=:mt"; params["mt"] = mapped_type
    if unmapped_sku:
        q += " AND (ps.uniware_sku IS NULL OR ps.uniware_sku='')"
    if search:
        q += " AND (ps.sku LIKE :s OR ps.order_id LIKE :s OR ps.item_description LIKE :s)"
        params["s"] = f"%{search}%"
    count_q = q.replace("SELECT ps.*, i.category", "SELECT COUNT(*)")
    total = db.execute(text(count_q), params).scalar()
    q += f" ORDER BY ps.invoice_date DESC LIMIT {per_page} OFFSET {(page-1)*per_page}"
    rows = db.execute(text(q), params).fetchall()
    return {"total": total, "page": page, "per_page": per_page,
            "items": [dict(r._mapping) for r in rows]}

@app.post("/api/portal-sales/map-sku/{sale_id}")
def map_portal_sku(sale_id: int, data: dict, db: Session = Depends(get_db), current_user=Depends(auth.get_current_user)):
    from models import SkuMapping
    portal_sku = data["portal_sku"]; item_sku = data["item_sku"]; portal = data["portal"]
    # Validate item master SKU exists
    valid = db.execute(text("SELECT id FROM items WHERE sku=:s"), {"s": item_sku}).fetchone()
    if not valid:
        raise HTTPException(400, f"SKU '{item_sku}' not found in Item Master")
    db.execute(text("UPDATE portal_sales SET uniware_sku=:u WHERE sku=:s AND portal=:p"),
               {"u": item_sku, "s": portal_sku, "p": portal})
    exists = db.execute(text("SELECT id FROM sku_mappings WHERE portal_sku=:s AND portal=:p"),
                        {"s": portal_sku, "p": portal}).fetchone()
    if exists:
        db.execute(text("UPDATE sku_mappings SET item_master_sku=:u WHERE portal_sku=:s AND portal=:p"),
                   {"u": item_sku, "s": portal_sku, "p": portal})
    else:
        db.add(SkuMapping(portal_sku=portal_sku, portal=portal, item_master_sku=item_sku))
    db.commit()
    return {"success": True}

@app.get("/api/portal-sales/summary")
def get_portal_sales_summary(
    portal: str = None, month: str = None, year: str = None,
    db: Session = Depends(get_db), current_user=Depends(auth.get_current_user)
):
    # Build WHERE clause and params cleanly
    conditions = ["1=1"]
    params = {}
    if portal and portal not in ("", "ALL"):
        conditions.append("portal=:portal"); params["portal"] = portal
    if month and month not in ("", "ALL"):
        conditions.append("month=:month"); params["month"] = month.upper().strip()
    if year and str(year) not in ("", "ALL", "0"):
        conditions.append("year=:year"); params["year"] = int(year)
    w = "WHERE " + " AND ".join(conditions)

    sale_row = db.execute(text(f"""
        SELECT COUNT(*), SUM(quantity), SUM(invoice_amount), SUM(tax_exclusive_gross)
        FROM portal_sales {w} AND mapped_type='Sale' AND quantity > 0
    """), params).fetchone()

    ret_row = db.execute(text(f"""
        SELECT COUNT(*), SUM(ABS(quantity)), SUM(ABS(invoice_amount)), SUM(ABS(tax_exclusive_gross))
        FROM portal_sales {w} AND mapped_type='Return'
    """), params).fetchone()

    unmapped = db.execute(text(f"SELECT COUNT(*) FROM portal_sales {w} AND mapped_type IS NULL"), params).scalar()
    unsku    = db.execute(text(f"SELECT COUNT(DISTINCT sku) FROM portal_sales {w} AND mapped_type='Sale' AND (uniware_sku IS NULL OR uniware_sku='')"), params).scalar()

    # Avg SP for unmapped SKUs — use tax_exclusive_gross (net of tax) for accurate SP
    urow = db.execute(text(f"""
        SELECT SUM(quantity), SUM(tax_exclusive_gross)
        FROM portal_sales {w} AND mapped_type='Sale'
        AND (uniware_sku IS NULL OR uniware_sku='') AND quantity > 0
    """), params).fetchone()
    uqty = urow[0] or 0
    uamt = urow[1] or 0

    sq = sale_row[1] or 0;  sa = sale_row[2] or 0;  steg = sale_row[3] or 0
    rq = ret_row[1]  or 0;  ra = ret_row[2]  or 0;  rteg = ret_row[3]  or 0
    nq = sq - rq
    na = round(sa - ra, 2)
    # avg_sp = sale taxable amount / sale qty (net of returns in amount, but qty from sales only for stability)
    # Use sale qty as denominator — meaningful even when returns dominate in filtered views
    avg_sp = round(steg / sq, 2) if sq > 0 else 0

    return {
        "sale_orders":       sale_row[0] or 0,
        "sale_qty":          sq,
        "sale_amt":          round(sa, 2),
        "sale_amount":       round(sa, 2),
        "return_orders":     ret_row[0] or 0,
        "return_qty":        rq,
        "return_amt":        round(ra, 2),
        "return_amount":     round(ra, 2),
        "net_qty":           nq,
        "net_amt":           na,
        "avg_sp":            avg_sp,
        "unmapped_types":    int(unmapped or 0),
        "unmapped_skus":     int(unsku or 0),
        "unmapped_qty":      int(uqty),
        "unmapped_amt":      round(uamt, 2),
        "unmapped_avg_sp":   round(uamt / uqty, 2) if uqty else 0,
    }

@app.get("/api/items/summary")
def get_items_summary(db: Session = Depends(get_db),
                      current_user=Depends(auth.get_current_user)):
    from sqlalchemy import text
    row = db.execute(text("""
        SELECT COUNT(*) as total,
               SUM(uniware_stock) as uniware,
               SUM(fba_stock) as fba,
               SUM(sjit_stock) as sjit,
               SUM(fbf_stock) as fbf,
               MAX(synced_at) as last_synced
        FROM items
    """)).fetchone()
    return {
        "total":       row[0] or 0,
        "uniware":     row[1] or 0,
        "fba":         row[2] or 0,
        "sjit":        row[3] or 0,
        "fbf":         row[4] or 0,
        "last_synced": row[5] or None,
    }

@app.get("/api/items")
def get_items(search: str = None, page: int = 1, per_page: int = 50,
              sort: str = "sku", dir: str = "asc",
              db: Session = Depends(get_db),
              current_user=Depends(auth.get_current_user)):
    from sqlalchemy import text
    # Whitelist sort columns to prevent SQL injection
    allowed = {"sku","category","style_id","uniware_stock","fba_stock","sjit_stock","fbf_stock","location","mrp","cost_price"}
    sort_col = sort if sort in allowed else "sku"
    sort_dir = "DESC" if dir == "desc" else "ASC"
    q = "SELECT * FROM items WHERE 1=1"
    params = {}
    if search:
        q += " AND (sku LIKE :s OR category LIKE :s OR catalog_name LIKE :s OR style_id LIKE :s OR location LIKE :s)"
        params["s"] = f"%{search}%"
    total = db.execute(text(q.replace("SELECT *", "SELECT COUNT(*)")), params).scalar()
    q += f" ORDER BY {sort_col} {sort_dir} LIMIT {per_page} OFFSET {(page-1)*per_page}"
    rows = db.execute(text(q), params).fetchall()
    return {"total": total, "page": page, "per_page": per_page,
            "items": [dict(r._mapping) for r in rows]}

@app.get("/api/items/sync-stream")
async def sync_items_stream(token: str, db: Session = Depends(get_db)):
    import json, csv, io, urllib.request
    from fastapi.responses import StreamingResponse
    from jose import jwt, JWTError
    from sqlalchemy import text
    from datetime import datetime
    SECRET_KEY = "rajnandini-fashion-secret-key-2024"
    try:
        payload = jwt.decode(token, SECRET_KEY, algorithms=["HS256"])
        if not payload.get("sub"):
            raise HTTPException(401, "Invalid token")
    except JWTError:
        raise HTTPException(401, "Invalid token")

    async def stream():
        try:
            _d = {"stage": "fetch", "pct": 10, "msg": "Fetching from Google Sheet..."}
            yield f"data: {json.dumps(_d)}\n\n"

            # Fetch CSV
            req = urllib.request.Request(SHEET_URL, headers={"User-Agent": "Mozilla/5.0"})
            try:
                with urllib.request.urlopen(req, timeout=20) as r:
                    content = r.read().decode("utf-8")
            except Exception as e:
                _d = {"stage": "error", "msg": f"Failed to fetch sheet: {str(e)}"}
                yield f"data: {json.dumps(_d)}\n\n"
                return

            _d = {"stage": "parse", "pct": 30, "msg": "Parsing sheet data..."}
            yield f"data: {json.dumps(_d)}\n\n"

            # Parse — find header row (row with "Uniware SKU")
            reader = csv.reader(io.StringIO(content))
            all_rows = list(reader)
            header_row_idx = None
            col_index = {}
            for i, row in enumerate(all_rows):
                if "Uniware SKU" in row:
                    header_row_idx = i
                    for j, h in enumerate(row):
                        h = h.strip()
                        if h in ITEM_COL_MAP:
                            col_index[ITEM_COL_MAP[h]] = j
                    break

            if header_row_idx is None or "sku" not in col_index:
                _d = {"stage": "error", "msg": "Could not find 'Uniware SKU' header in sheet"}
                yield f"data: {json.dumps(_d)}\n\n"
                return

            _d = {"stage": "parse", "pct": 50, "msg": f"Headers found at row {header_row_idx+1}. Syncing..."}
            yield f"data: {json.dumps(_d)}\n\n"

            # Sync rows
            def _v(row, field):
                idx = col_index.get(field)
                if idx is None or idx >= len(row): return ""
                return row[idx].strip()

            def _f(row, field):
                try: return float(_v(row, field).replace(",", "") or 0)
                except: return 0.0

            def _i(row, field):
                try: return int(float(_v(row, field).replace(",", "") or 0))
                except: return 0

            added = updated = skipped = 0
            now = datetime.now()

            for row in all_rows[header_row_idx + 1:]:
                sku = _v(row, "sku")
                if not sku:
                    skipped += 1
                    continue
                data = {
                    "sku":           sku,
                    "category":      _v(row, "category"),
                    "style_id":      _v(row, "style_id"),
                    "availability":  _v(row, "availability"),
                    "uniware_stock": _i(row, "uniware_stock"),
                    "fba_stock":     _i(row, "fba_stock"),
                    "sjit_stock":    _i(row, "sjit_stock"),
                    "fbf_stock":     _i(row, "fbf_stock"),
                    "location":      _v(row, "location"),
                    "remark":        _v(row, "remark"),
                    "cost_price":    _f(row, "cost_price"),
                    "mrp":           _f(row, "mrp"),
                    "catalog_name":  _v(row, "catalog_name"),
                    "synced_at":     now,
                }
                existing = db.execute(text("SELECT id FROM items WHERE sku=:sku"), {"sku": sku}).fetchone()
                if existing:
                    sets = ", ".join(f"{k}=:{k}" for k in data if k != "sku")
                    db.execute(text(f"UPDATE items SET {sets} WHERE sku=:sku"), data)
                    updated += 1
                else:
                    cols = ", ".join(data.keys())
                    vals = ", ".join(f":{k}" for k in data.keys())
                    db.execute(text(f"INSERT INTO items ({cols}) VALUES ({vals})"), data)
                    added += 1

            db.commit()
            total = added + updated
            _d = {"stage": "done", "pct": 100, "added": added, "updated": updated,
                  "skipped": skipped, "total": total,
                  "msg": f"Sync complete! {added} added, {updated} updated"}
            yield f"data: {json.dumps(_d)}\n\n"

        except Exception as e:
            import traceback
            _d = {"stage": "error", "msg": str(e) + " | " + traceback.format_exc()[-300:]}
            yield f"data: {json.dumps(_d)}\n\n"

    return StreamingResponse(stream(), media_type="text/event-stream",
        headers={"Cache-Control": "no-cache", "X-Accel-Buffering": "no"})

# ── SKU MAPPING ───────────────────────────────────────────────────────────────

@app.get("/api/sku-mappings")
def get_sku_mappings(portal: str = None, unmapped_only: bool = False,
                     db: Session = Depends(get_db),
                     current_user=Depends(auth.get_current_user)):
    from sqlalchemy import text
    q = "SELECT * FROM sku_mappings WHERE 1=1"
    params = {}
    if portal:
        q += " AND portal = :portal"
        params["portal"] = portal
    if unmapped_only:
        q += " AND (item_master_sku IS NULL OR item_master_sku = '')"
    q += " ORDER BY portal, portal_sku"
    rows = db.execute(text(q), params).fetchall()
    return [dict(r._mapping) for r in rows]

@app.put("/api/sku-mappings/{mapping_id}")
def update_sku_mapping(mapping_id: int, data: dict,
                       db: Session = Depends(get_db),
                       current_user=Depends(auth.get_current_user)):
    item_sku = (data.get("item_master_sku") or "").strip()
    if item_sku:
        # Validate item exists in item master
        exists = db.execute(text("SELECT id FROM items WHERE sku=:s"), {"s": item_sku}).fetchone()
        if not exists:
            raise HTTPException(400, f"SKU '{item_sku}' not found in Item Master")
    db.execute(text("UPDATE sku_mappings SET item_master_sku=:sku, updated_at=CURRENT_TIMESTAMP WHERE id=:id"),
               {"sku": item_sku or None, "id": mapping_id})
    # Also update portal_sales
    mapping = db.execute(text("SELECT portal_sku, portal FROM sku_mappings WHERE id=:id"), {"id": mapping_id}).fetchone()
    if mapping and item_sku:
        db.execute(text("UPDATE portal_sales SET uniware_sku=:u WHERE sku=:s AND portal=:p"),
                   {"u": item_sku, "s": mapping.portal_sku, "p": mapping.portal})
    db.commit()
    return {"ok": True}

@app.delete("/api/sku-mappings/{mapping_id}")
def delete_sku_mapping(mapping_id: int, db: Session = Depends(get_db),
                       current_user=Depends(auth.get_current_user)):
    from sqlalchemy import text
    db.execute(text("DELETE FROM sku_mappings WHERE id = :id"), {"id": mapping_id})
    db.commit()
    return {"ok": True}

@app.post("/api/sku-mappings/import")
async def import_sku_mappings(file: UploadFile = File(...),
                              db: Session = Depends(get_db),
                              current_user=Depends(auth.get_current_user)):
    import csv, io
    content = await file.read()
    reader = csv.DictReader(io.StringIO(content.decode("utf-8-sig")))
    rows = list(reader)
    # Get all valid item master SKUs
    valid_skus = {r.sku for r in db.execute(text("SELECT sku FROM items")).fetchall()}
    inserted = 0; updated = 0; skipped = 0
    for row in rows:
        portal_sku = (row.get("portal_sku") or "").strip()
        portal     = (row.get("portal") or "").strip()
        item_sku   = (row.get("item_master_sku") or "").strip()
        if not portal_sku or not portal:
            skipped += 1; continue
        # Validate item master SKU if provided
        if item_sku and item_sku not in valid_skus:
            skipped += 1; continue
        existing = db.execute(text("SELECT id FROM sku_mappings WHERE portal_sku=:ps AND portal=:p"),
                              {"ps": portal_sku, "p": portal}).fetchone()
        if existing:
            db.execute(text("UPDATE sku_mappings SET item_master_sku=:sku, updated_at=CURRENT_TIMESTAMP WHERE id=:id"),
                      {"sku": item_sku or None, "id": existing[0]})
            updated += 1
        else:
            db.execute(text("INSERT INTO sku_mappings (portal_sku, portal, item_master_sku) VALUES (:ps, :p, :sku)"),
                      {"ps": portal_sku, "p": portal, "sku": item_sku or None})
            inserted += 1
        # Also update portal_sales
        if item_sku:
            db.execute(text("UPDATE portal_sales SET uniware_sku=:u WHERE sku=:ps AND portal=:p"),
                       {"u": item_sku, "ps": portal_sku, "p": portal})
    db.commit()
    return {"inserted": inserted, "updated": updated, "skipped": skipped, "total": len(rows)}

@app.get("/api/sku-mappings/export")
def export_sku_mappings(unmapped_only: bool = False, token: str = None,
                        db: Session = Depends(get_db)):
    import csv, io
    from fastapi.responses import StreamingResponse
    from jose import jwt
    SECRET_KEY = "rajnandini-fashion-secret-key-2024"
    try:
        payload = jwt.decode(token, SECRET_KEY, algorithms=["HS256"])
        if not payload.get("sub"): raise HTTPException(401)
    except: raise HTTPException(401, "Not authenticated")
    q = "SELECT portal_sku, portal, item_master_sku FROM sku_mappings"
    if unmapped_only:
        q += " WHERE item_master_sku IS NULL OR item_master_sku = ''"
    q += " ORDER BY portal, portal_sku"
    rows = db.execute(text(q)).fetchall()
    output = io.StringIO()
    w = csv.writer(output)
    w.writerow(["portal_sku", "portal", "item_master_sku"])
    for r in rows:
        w.writerow([r[0], r[1], r[2] or ""])
    output.seek(0)
    return StreamingResponse(iter([output.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": "attachment; filename=sku_mappings.csv"})



# ─── ALERTS ──────────────────────────────────────────────────────────────────

@app.get("/api/alerts/recon")
def get_recon_alerts(portal: str = "amazon", month: str = None, year: str = None,
                     db: Session = Depends(get_db), current_user=Depends(auth.get_current_user)):
    return {"unmapped_skus": 0, "missing_cost": 0}

@app.get("/api/alerts/pending")
def get_pending_alerts(db: Session = Depends(get_db), current_user=Depends(auth.get_current_user)):
    unmapped = db.execute(text("""SELECT COUNT(DISTINCT ps.sku) FROM portal_sales ps
        WHERE ps.mapped_type='Sale' AND (ps.uniware_sku IS NULL OR ps.uniware_sku='')""")).scalar() or 0
    missing_cost = db.execute(text("""SELECT COUNT(DISTINCT ps.sku) FROM portal_sales ps
        LEFT JOIN sku_mappings sm ON ps.sku=sm.portal_sku AND ps.portal=sm.portal
        LEFT JOIN items i ON (sm.item_master_sku=i.sku OR ps.sku=i.sku)
        WHERE ps.mapped_type='Sale'
        AND (ps.uniware_sku IS NOT NULL AND ps.uniware_sku!='')
        AND (i.cost_price IS NULL OR i.cost_price=0)""")).scalar() or 0
    return {"unmapped_skus": unmapped, "missing_cost": missing_cost}


# ── PHYSICAL RETURN ───────────────────────────────────────────────────────────
@app.post("/api/physical-returns/sync")
def sync_physical_returns(
    data: schemas.PhysicalReturnSync,
    db: Session = Depends(get_db),
    current_user=Depends(auth.get_current_user)
):
    rows = [r.dict() for r in data.rows]
    return crud.sync_physical_returns(db, rows, replace=data.replace)

@app.get("/api/physical-returns")
def get_physical_returns(
    db: Session = Depends(get_db),
    current_user=Depends(auth.get_current_user)
):
    rows = crud.get_physical_returns(db)
    return [
        {
            "date":         r.date,
            "channel":      r.channel,
            "order_no":     r.order_no,
            "awb":          r.awb,
            "courier":      r.courier,
            "putway":       r.putway,
            "sku_r":        r.sku_r,
            "to_rma":       r.to_rma,
            "remark":       r.remark,
            "putaway_code": r.putaway_code,
        }
        for r in rows
    ]

@app.get("/", response_class=HTMLResponse)
def index(request: Request):
    return templates.TemplateResponse("index.html", {"request": request})

if __name__ == "__main__":
    uvicorn.run("main:app", host="0.0.0.0", port=8000, reload=True)
